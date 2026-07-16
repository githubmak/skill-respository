"""
Split Storyboard To Script — XLSX Generator

Generates a formatted xlsx workbook from script/storyboard analysis data.
Requires the user-confirmed final export directory; no implicit output path.

Sheets:
  1. 视觉规范表  — Scene visual spec (characters, scenes, lighting, global params)
  2. 分镜表      — Shot-by-shot table with all fields
  3. 子画面表    — Subframe/keyframe details
  4. 单镜提示词  — Per-shot structured prompts (fixed prefix + dynamic + negatives)
  5. 完整拼接    — Reviewed full prompt ready for AI tools
"""

import json
import re
from pathlib import Path
from datetime import datetime


def _clean_text(value: object) -> str:
    return str(value or '').strip()


def _join_segments(parts: list[str], separator: str = '，') -> str:
    cleaned = [part.strip('，；; ') for part in parts if _clean_text(part)]
    return separator.join(cleaned)


def _normalize_visual_prompt_text(text: object) -> str:
    result = _clean_text(text)
    if not result:
        return ''

    result = re.sub(r'【[^】]+】', '', result)
    result = re.sub(r'[ \t]+', ' ', result)
    result = re.sub(r'\n{3,}', '\n\n', result)
    lines = [line.strip('，；; ') for line in result.splitlines()]
    lines = [line for line in lines if line]
    return '\n'.join(lines)


def _normalize_subframes(shot: dict) -> list[dict]:
    """Normalize optional subframe data for one master shot."""
    raw_subframes = shot.get('subframes', []) or []
    normalized = []
    shot_id = shot.get('id', 'S1-00')
    for idx, sub in enumerate(raw_subframes, 1):
        sub_id = sub.get('id') or f'{shot_id}-{idx}'
        normalized.append({
            'id': sub_id,
            'label': sub.get('label', f'子画面{idx}'),
            'frame_role': sub.get('frame_role', ''),
            'time_start': sub.get('time_start', ''),
            'time_end': sub.get('time_end', ''),
            'transition_in': sub.get('transition_in', ''),
            'is_key_frame': bool(sub.get('is_key_frame', True)),
            'shot_size': sub.get('shot_size', ''),
            'camera_movement': sub.get('camera_movement', sub.get('camera', '')),
            'camera_angle': sub.get('camera_angle', ''),
            'lens_spec': sub.get('lens_spec', ''),
            'composition': sub.get('composition', ''),
            'expression': sub.get('expression', ''),
            'action': sub.get('action', ''),
            'dialogue': sub.get('dialogue', ''),
            'lighting': sub.get('lighting', ''),
            'continuity_anchor': sub.get('continuity_anchor', ''),
        })
    return normalized


def _subframes_summary(subframes: list[dict]) -> str:
    if not subframes:
        return ''
    parts = []
    for sub in subframes:
        key_tag = '关键画面' if sub.get('is_key_frame') else '辅助画面'
        summary = sub.get('composition') or sub.get('action') or sub.get('dialogue') or sub.get('expression') or sub.get('label', '')
        parts.append(f"{sub.get('id')}[{key_tag}]：{summary}")
    return '\n'.join(parts)


def _subframe_camera_prompt(sub: dict) -> str:
    """Build a camera-first phrase for one subframe."""
    lens_spec = (sub.get('lens_spec') or '').strip()
    shot_size = (sub.get('shot_size') or '').strip()
    angle = (sub.get('camera_angle') or '').strip()
    movement = (sub.get('camera_movement') or '').strip()

    if not lens_spec and shot_size:
        lens_spec = _infer_lens_spec(shot_size)
    if not angle:
        angle = _infer_camera_angle(sub.get('composition', ''))

    parts = []
    if lens_spec:
        parts.append(lens_spec)
    elif shot_size:
        parts.append(shot_size)
    if angle:
        parts.append(angle)
    if movement:
        parts.append(movement)
    return '\uff0c'.join(parts)
    return '，'.join(parts)


def _subframes_prompt(subframes: list[dict]) -> str:
    if not subframes:
        return ''
    lines = [f'\u3010\u5b50\u753b\u9762\u3011\uff08\u5171{len(subframes)}\u4e2a\uff0c\u5747\u4e3a\u672c\u955c\u5934\u5173\u952e\u753b\u9762\u652f\u6491\uff09']
    for sub in subframes:
        key_tag = '\u5173\u952e\u753b\u9762' if sub.get('is_key_frame') else '\u8f85\u52a9\u753b\u9762'
        detail_parts = []
        camera_prompt = _subframe_camera_prompt(sub)
        if camera_prompt:
            detail_parts.append(f"\u955c\u5934\uff1a{camera_prompt}")
        if sub.get('composition'):
            detail_parts.append(f"\u753b\u9762\uff1a{sub['composition']}")
        if sub.get('expression'):
            detail_parts.append(f"\u60c5\u7eea\uff1a{sub['expression']}")
        if sub.get('action'):
            detail_parts.append(f"\u52a8\u4f5c\uff1a{sub['action']}")
        if sub.get('dialogue'):
            detail_parts.append(f"\u53f0\u8bcd\uff1a{sub['dialogue']}")
        if sub.get('lighting'):
            detail_parts.append(f"\u5149\u5f71\uff1a{sub['lighting']}")
        if sub.get('continuity_anchor'):
            detail_parts.append(f"\u4e00\u81f4\u6027\u951a\u70b9\uff1a{sub['continuity_anchor']}")
        time_start = sub.get('time_start', '')
        time_end = sub.get('time_end', '')
        time_tag = f"{time_start}-{time_end}秒｜" if time_start != '' and time_end != '' else ''
        transition = f"转换：{sub.get('transition_in')}；" if sub.get('transition_in') else ''
        lines.append(f"\u25b8 {time_tag}{sub.get('id')}\uff08{key_tag}\uff09\uff1a{transition}" + '\uff1b'.join(detail_parts))
    return '\n'.join(lines)
    lines = [f'【子画面】（共{len(subframes)}个，均为本镜头关键画面支撑）']
    for sub in subframes:
        key_tag = '关键画面' if sub.get('is_key_frame') else '辅助画面'
        detail_parts = []
        camera_prompt = _subframe_camera_prompt(sub)
        if camera_prompt:
            detail_parts.append(f"镜头：{camera_prompt}")
        if sub.get('composition'):
            detail_parts.append(f"画面：{sub['composition']}")
        if sub.get('expression'):
            detail_parts.append(f"情绪：{sub['expression']}")
        if sub.get('action'):
            detail_parts.append(f"动作：{sub['action']}")
        if sub.get('dialogue'):
            detail_parts.append(f"台词：{sub['dialogue']}")
        if sub.get('lighting'):
            detail_parts.append(f"光影：{sub['lighting']}")
        if sub.get('continuity_anchor'):
            detail_parts.append(f"一致性锚点：{sub['continuity_anchor']}")
        lines.append(f"▸ {sub.get('id')}（{key_tag}）：" + '；'.join(detail_parts))
    return '\n'.join(lines)
    lines = [f'【子画面】（共{len(subframes)}个，均为本镜头关键画面支撑）']
    for sub in subframes:
        key_tag = '关键画面' if sub.get('is_key_frame') else '辅助画面'
        detail_parts = []
        if sub.get('composition'):
            detail_parts.append(f"画面：{sub['composition']}")
        if sub.get('expression'):
            detail_parts.append(f"情绪：{sub['expression']}")
        if sub.get('action'):
            detail_parts.append(f"动作：{sub['action']}")
        if sub.get('dialogue'):
            detail_parts.append(f"台词：{sub['dialogue']}")
        if sub.get('lighting'):
            detail_parts.append(f"光影：{sub['lighting']}")
        if sub.get('continuity_anchor'):
            detail_parts.append(f"一致性锚点：{sub['continuity_anchor']}")
        lines.append(f"▸ {sub.get('id')}（{key_tag}）：" + '；'.join(detail_parts))
    return '\n'.join(lines)


def _infer_lens_spec(shot_size: str) -> str:
    size = shot_size or ''
    if '大特写' in size:
        return '微距 100mm 大特写'
    if '特写' in size:
        return '85mm 长焦特写'
    if '近景' in size and '中近景' not in size:
        return '85mm 长焦近景'
    if '中近景' in size:
        return '85mm 长焦中近景'
    if '中景' in size:
        return '50mm 中景'
    if '中全景' in size:
        return '50mm 中全景'
    if '全景' in size:
        return '35mm 全景'
    if '远景' in size:
        return '24mm 远景'
    return size or '50mm 镜头'


def _infer_camera_angle(text: str) -> str:
    source = text or ''
    if '俯' in source:
        return '俯视'
    if '仰' in source:
        return '仰视'
    if '侧' in source:
        return '侧视'
    return '平视'


def _subframes_inline(subframes: list[dict]) -> str:
    if not subframes:
        return ''
    lines = []
    for idx, sub in enumerate(subframes, 1):
        camera_prompt = _subframe_camera_prompt(sub) or '\u955c\u5934\u4fdd\u6301\u9759\u6b62'
        pieces = [
            f'\u955c\u5934 {idx}\uff1a{camera_prompt}',
            sub.get('composition', ''),
            sub.get('expression', ''),
            sub.get('action', ''),
            sub.get('dialogue', ''),
            sub.get('lighting', ''),
        ]
        line = '\uff1b'.join(part.strip('\uff0c\uff1b ') for part in pieces if part)
        lines.append(line)
    return '\n'.join(lines)
    lines = []
    for idx, sub in enumerate(subframes, 1):
        camera_prompt = _subframe_camera_prompt(sub) or '镜头保持静止'
        pieces = [
            f'镜头 {idx}：{camera_prompt}',
            sub.get('composition', ''),
            sub.get('expression', ''),
            sub.get('action', ''),
            sub.get('dialogue', ''),
            sub.get('lighting', ''),
        ]
        line = '；'.join(part.strip('，； ') for part in pieces if part)
        lines.append(line)
    return '\n'.join(lines)
    lines = []
    for idx, sub in enumerate(subframes, 1):
        lens_spec = sub.get('lens_spec') or _infer_lens_spec(sub.get('shot_size', ''))
        angle = sub.get('camera_angle') or _infer_camera_angle(sub.get('composition', ''))
        movement = sub.get('camera_movement', '') or '镜头保持静止'
        pieces = [
            f'镜头 {idx}：{angle} {lens_spec}',
            movement,
            sub.get('composition', ''),
            sub.get('expression', ''),
            sub.get('action', ''),
            sub.get('dialogue', ''),
            sub.get('lighting', ''),
        ]
        line = '，'.join(part.strip('，； ') for part in pieces if part)
        lines.append(line)
    return '\n'.join(lines)


def _enrich_shot_with_subframes(shot: dict) -> dict:
    enriched = dict(shot)
    subframes = _normalize_subframes(shot)
    validation = shot.get('validation', {}) or {}
    enriched['subframes'] = subframes
    enriched['subframe_count'] = len(subframes)
    enriched['subframes_summary'] = _subframes_summary(subframes)
    enriched['subframes_prompt'] = _subframes_prompt(subframes)
    enriched['subframes_inline'] = _subframes_inline(subframes)
    enriched['validation_status'] = (
        shot.get('validation_status')
        or validation.get('status')
        or ('通过' if shot.get('validation_passed') is True else '')
    )
    enriched['validation_notes'] = shot.get('validation_notes') or validation.get('notes', '')
    return enriched


def _compose_full_prompt(
    shot: dict,
    scene_fixed_prompt: str = '',
    unified_negatives: str = '',
) -> str:
    """Prefer the reviewed full prompt; otherwise build from structured fields."""
    if shot.get('full_prompt'):
        return _normalize_visual_prompt_text(shot['full_prompt'])

    shot_scene_fixed = shot.get('scene_fixed_prefix', scene_fixed_prompt)
    shot_dynamic = shot.get('shot_dynamic_prompt', '')
    shot_subframes = shot.get('subframes_prompt', '')
    shot_negatives = shot.get('unified_negatives', unified_negatives)

    parts = [part for part in [shot_scene_fixed, shot_dynamic, shot_subframes, shot_negatives] if part]
    return _normalize_visual_prompt_text('\n\n'.join(parts))


def _is_validation_failed(shot: dict) -> bool:
    status = str(shot.get('validation_status', '')).strip()
    return any(flag in status for flag in ('未通过', '不通过', '失败', 'fail', 'failed'))


def _build_visual_subframes_prompt(subframes: list[dict]) -> str:
    if not subframes:
        return ''

    lines = [f'【关键画面】（共{len(subframes)}个，用于稳定人物关系、目标物件和情绪推进）']
    for sub in subframes:
        key_tag = '关键画面' if sub.get('is_key_frame') else '辅助画面'
        parts = []
        if sub.get('composition'):
            parts.append(f"画面：{_normalize_visual_prompt_text(sub['composition'])}")
        if sub.get('expression'):
            parts.append(f"情绪：{_normalize_visual_prompt_text(sub['expression'])}")
        if sub.get('action'):
            parts.append(f"人物状态：{_normalize_visual_prompt_text(sub['action'])}")
        if sub.get('dialogue'):
            parts.append(f"声音：{_normalize_visual_prompt_text(sub['dialogue'])}")
        if sub.get('lighting'):
            parts.append(f"光影：{_normalize_visual_prompt_text(sub['lighting'])}")
        if sub.get('continuity_anchor'):
            parts.append(f"一致性锚点：{_normalize_visual_prompt_text(sub['continuity_anchor'])}")
        lines.append(f"▸ {sub.get('id')}（{key_tag}）：" + '；'.join(parts))
    return '\n'.join(lines)


def _build_visual_subframes_inline(subframes: list[dict]) -> str:
    if not subframes:
        return ''

    lines = []
    for idx, sub in enumerate(subframes, 1):
        time_start = sub.get('time_start', '')
        time_end = sub.get('time_end', '')
        time_tag = f'{time_start}-{time_end}秒' if time_start != '' and time_end != '' else ''
        parts = [
            f'关键画面 {idx}',
            time_tag,
            _normalize_visual_prompt_text(sub.get('transition_in', '')),
            _normalize_visual_prompt_text(sub.get('shot_size', '')),
            _normalize_visual_prompt_text(sub.get('camera_movement', '')),
            _normalize_visual_prompt_text(sub.get('composition', '')),
            _normalize_visual_prompt_text(sub.get('expression', '')),
            _normalize_visual_prompt_text(sub.get('action', '')),
            _normalize_visual_prompt_text(sub.get('dialogue', '')),
            _normalize_visual_prompt_text(sub.get('lighting', '')),
        ]
        lines.append('；'.join(part.strip('，； ') for part in parts if part))
    return '\n'.join(lines)


def _build_visual_dynamic_prompt(shot: dict) -> str:
    parts = []

    if shot.get('characters_present'):
        parts.append(f"出场人物：{_normalize_visual_prompt_text(shot['characters_present'])}")
    if shot.get('organization_mode'):
        parts.append(f"组织模式：{_normalize_visual_prompt_text(shot['organization_mode'])}")
    if shot.get('duration'):
        parts.append(f"总时长：{shot['duration']}秒")

    if shot.get('composition'):
        parts.append(f"画面：{_normalize_visual_prompt_text(shot['composition'])}")
    if shot.get('expression'):
        parts.append(f"情绪：{_normalize_visual_prompt_text(shot['expression'])}")

    state_text = _join_segments([
        _normalize_visual_prompt_text(shot.get('action', '')),
        _normalize_visual_prompt_text(shot.get('interaction', '')),
    ], '；')
    if state_text:
        parts.append(f"人物状态：{state_text}")

    sound_text = _join_segments([
        _normalize_visual_prompt_text(shot.get('dialogue', '')),
        _normalize_visual_prompt_text(shot.get('audio', '')),
    ], '；')
    if sound_text:
        parts.append(f"声音：{sound_text}")

    detail_text = _join_segments([
        _normalize_visual_prompt_text(shot.get('lighting_enhancement', '')),
        _normalize_visual_prompt_text(shot.get('rhythm', '')),
    ], '；')
    if detail_text:
        parts.append(f"氛围细节：{detail_text}")

    return '\n'.join(parts)


def _is_validation_approved(shot: dict) -> bool:
    status = str(shot.get('validation_status', '')).strip().lower()
    if _is_validation_failed(shot):
        return False

    pass_tokens = ('通过', 'pass', 'passed', 'approved')
    has_pass = any(token in status for token in pass_tokens)
    has_notes = bool(str(shot.get('validation_notes', '')).strip())
    has_reviewed_prompt = bool(str(shot.get('full_prompt', '')).strip())
    return has_pass and has_notes and has_reviewed_prompt


def generate_xlsx(
    data: dict,
    output_dir: Path | None = None,
    filename_prefix: str | None = None,
) -> Path:
    """
    Generate the structured prompt workbook.

    Args:
        data: Full analysis data dict with keys:
            - meta: {script_title, aspect_ratio, style, render_engine, max_duration, ...}
            - characters: [{name, age, face, features, hair, costume, build, voice, ...}]
            - scenes: [{name, indoor_outdoor, props, atmosphere, time_of_day, ...}]
            - lighting: {source_direction, color_temp, contrast, fill_light, ...}
            - scene_fixed_prompt: str  (the scene-level fixed prefix)
            - unified_negatives: str
            - shots: [{id, scene, duration, shot_size, camera_movement, composition,
                       expression, action, interaction, dialogue, audio, rhythm,
                       lighting_enhancement, special_negatives, ...}]
        output_dir: User-confirmed final export directory. Required; no implicit fallback.
        filename_prefix: Optional custom prefix for the filename

    Returns:
        Path to the generated xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl is required. Install: pip install openpyxl")
        raise

    wb = Workbook()

    meta = data.get('meta', {})
    characters = data.get('characters', [])
    scenes = data.get('scenes', [])
    lighting = data.get('lighting', {})
    scene_fixed_prompt = data.get('scene_fixed_prompt', '')
    unified_negatives = data.get('unified_negatives', '')
    shots = [_enrich_shot_with_subframes(shot) for shot in data.get('shots', [])]
    special_negative_types = data.get('special_negative_types', {})

    script_title = meta.get('script_title', '未命名剧本')
    aspect_ratio = meta.get('aspect_ratio', '16:9横屏')
    style_name = meta.get('style', '')
    render_engine = meta.get('render_engine', '')
    max_duration = meta.get('max_duration', 5)
    genre = meta.get('genre', '')

    # ── Styles ──
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    section_font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')

    body_font = Font(name='微软雅黑', size=10)
    body_align = Alignment(vertical='top', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

    title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
    subtitle_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
    meta_font = Font(name='微软雅黑', size=10, color='555555')

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    header_border = Border(
        left=Side(style='thin', color='1F3864'),
        right=Side(style='thin', color='1F3864'),
        top=Side(style='thin', color='1F3864'),
        bottom=Side(style='thin', color='1F3864'),
    )

    # Tier fills
    tier_fills = {
        'S': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'A': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'B': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    }

    highlight_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

    # Stats
    total_dur = sum(float(s.get('duration', 3)) for s in shots)
    total_subframes = sum(int(s.get('subframe_count', 0)) for s in shots)
    mins, secs = int(total_dur // 60), int(total_dur % 60)
    dur_str = f'~{mins}分{secs}秒' if mins > 0 else f'~{secs}秒'
    avg_dur = round(total_dur / len(shots), 1) if shots else 0

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: 视觉规范表
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '视觉规范表'

    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = f'{script_title} — 本场视觉规范表'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 28

    # Meta summary
    ws1.merge_cells('A2:F2')
    ws1['A2'] = (
        f'画幅: {aspect_ratio} | 风格: {style_name} | 渲染引擎: {render_engine} | '
        f'题材: {genre} | 时长上限: {max_duration}s | '
        f'总镜数: {len(shots)} | 子画面: {total_subframes} | 总时长: {dur_str} | 均镜: ~{avg_dur}s | '
        f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )
    ws1['A2'].font = meta_font
    ws1['A2'].alignment = Alignment(vertical='center')
    ws1.row_dimensions[2].height = 22

    # ── Section: Global Config ──
    row = 4
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎全局配置').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    config_items = [
        ('画幅比例', aspect_ratio),
        ('渲染风格', render_engine),
        ('内容风格', style_name),
        ('内容题材', genre),
        ('单镜时长上限', f'{max_duration}秒'),
        ('时代背景', meta.get('era', '')),
    ]
    for j, (label, value) in enumerate(config_items):
        r = row + 1 + j
        ws1.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', bold=True, size=10)
        ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws1.cell(row=r, column=1).border = thin_border
        ws1.merge_cells(f'B{r}:F{r}')
        ws1.cell(row=r, column=2, value=value).font = body_font
        ws1.cell(row=r, column=2).alignment = body_align
        ws1.cell(row=r, column=2).border = thin_border

    # ── Section: Characters ──
    row = row + len(config_items) + 2
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎人物库').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    char_headers = ['姓名', '年龄', '脸型/五官', '发型/发饰', '服装', '身形气质', '声线']
    for col, h in enumerate(char_headers, 1):
        cell = ws1.cell(row=row + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws1.row_dimensions[row + 1].height = 22

    for i, c in enumerate(characters):
        r = row + 2 + i
        vals = [
            c.get('name', ''),
            c.get('age', ''),
            c.get('face_features', c.get('features', '')),
            c.get('hair', ''),
            c.get('costume', ''),
            c.get('build', ''),
            c.get('voice', ''),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=col, value=v)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
        ws1.row_dimensions[r].height = 36

    # ── Section: Scenes ──
    row = row + 2 + len(characters) + 1
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎场景库').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    scene_headers = ['场景名', '室内/室外', '时间', '核心陈设', '整体环境', '备注']
    for col, h in enumerate(scene_headers, 1):
        cell = ws1.cell(row=row + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws1.row_dimensions[row + 1].height = 22

    for i, sc in enumerate(scenes):
        r = row + 2 + i
        vals = [
            sc.get('name', ''),
            sc.get('indoor_outdoor', ''),
            sc.get('time_of_day', ''),
            sc.get('props', ''),
            sc.get('atmosphere', ''),
            sc.get('notes', ''),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=col, value=v)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
        ws1.row_dimensions[r].height = 32

    # ── Section: Lighting ──
    row = row + 2 + len(scenes) + 1
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎光影基调').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    light_items = [
        ('主光源方向', lighting.get('source_direction', '')),
        ('色温色调', lighting.get('color_temp', '')),
        ('明暗对比', lighting.get('contrast', '')),
        ('辅光', lighting.get('fill_light', '')),
        ('全局光影效果', lighting.get('effects', '')),
    ]
    for j, (label, value) in enumerate(light_items):
        r = row + 1 + j
        ws1.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', bold=True, size=10)
        ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws1.cell(row=r, column=1).border = thin_border
        ws1.merge_cells(f'B{r}:F{r}')
        ws1.cell(row=r, column=2, value=value).font = body_font
        ws1.cell(row=r, column=2).alignment = body_align
        ws1.cell(row=r, column=2).border = thin_border

    # ── Section: Scene-Level Fixed Prompt ──
    row = row + len(light_items) + 2
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎场级统一固定提示词').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    r = row + 1
    ws1.merge_cells(f'A{r}:F{r}')
    ws1.cell(row=r, column=1, value=scene_fixed_prompt).font = body_font
    ws1.cell(row=r, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws1.cell(row=r, column=1).fill = highlight_fill
    ws1.cell(row=r, column=1).border = thin_border
    ws1.row_dimensions[r].height = 72

    # ── Section: Unified Negatives ──
    row = r + 2
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value='▎统一负面提示词').font = section_font
    ws1.cell(row=row, column=1).fill = section_fill
    ws1.row_dimensions[row].height = 22

    r = row + 1
    ws1.merge_cells(f'A{r}:F{r}')
    ws1.cell(row=r, column=1, value=unified_negatives).font = body_font
    ws1.cell(row=r, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws1.cell(row=r, column=1).border = thin_border
    ws1.row_dimensions[r].height = 48

    # Special negative type table
    if special_negative_types:
        r2 = r + 2
        ws1.merge_cells(f'A{r2}:F{r2}')
        ws1.cell(row=r2, column=1, value='按镜头类型追加负面词').font = Font(name='微软雅黑', bold=True, size=10)
        ws1.cell(row=r2, column=1).fill = section_fill
        ws1.row_dimensions[r2].height = 20

        for i, (shot_type, neg_words) in enumerate(special_negative_types.items()):
            rr = r2 + 1 + i
            ws1.cell(row=rr, column=1, value=shot_type).font = Font(name='微软雅黑', bold=True, size=10)
            ws1.cell(row=rr, column=1).alignment = body_align
            ws1.cell(row=rr, column=1).border = thin_border
            ws1.merge_cells(f'B{rr}:F{rr}')
            ws1.cell(row=rr, column=2, value=neg_words).font = body_font
            ws1.cell(row=rr, column=2).alignment = body_align
            ws1.cell(row=rr, column=2).border = thin_border

    # Column widths
    for col, w in enumerate([14, 22, 22, 22, 22, 22], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: 分镜表
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('分镜表')

    SHOT_COLUMNS = [
        ('镜号', 'id', 12),
        ('场景', 'scene', 14),
        ('时长(s)', 'duration', 8),
        ('组织模式', 'organization_mode', 18),
        ('出场人物', 'characters_present', 24),
        ('层级', 'tier', 6),
        ('景别', 'shot_size', 20),
        ('运镜', 'camera_movement', 32),
        ('子画面数', 'subframe_count', 8),
        ('画面构图', 'composition', 40),
        ('人物神态', 'expression', 40),
        ('动作交互', 'action', 40),
        ('台词音频(含时长校验)', 'dialogue', 45),
        ('运镜节奏', 'rhythm', 30),
        ('光影强化', 'lighting_enhancement', 30),
        ('专项负面', 'special_negatives', 24),
        ('子画面摘要', 'subframes_summary', 55),
        ('子分镜', 'subframes_inline', 80),
        ('分镜图(起幅/落幅)', 'storyboard_frames_text', 50),
        ('校验状态', 'validation_status', 12),
        ('校验备注', 'validation_notes', 50),
    ]

    # Raise row height to accommodate storyboard frames text
    SHOT_ROW_HEIGHT = 80

    for col_idx, (header, _, width) in enumerate(SHOT_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws2.row_dimensions[1].height = 25
    ws2.freeze_panes = 'B2'

    for col_idx, (_, _, width) in enumerate(SHOT_COLUMNS, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, shot in enumerate(shots):
        r = row_idx + 2
        tier = shot.get('tier', '')
        row_fill = tier_fills.get(tier, None)

        for col_idx, (_, key, _) in enumerate(SHOT_COLUMNS, 1):
            val = shot.get(key, '') or ''
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
        ws2.row_dimensions[r].height = SHOT_ROW_HEIGHT

    if shots:
        ws2.auto_filter.ref = f'A1:{get_column_letter(len(SHOT_COLUMNS))}{len(shots) + 1}'

    # ═══════════════════════════════════════════════════════════════
    # Sheet 3: 子画面表
    # ═══════════════════════════════════════════════════════════════
    ws_sub = wb.create_sheet('子画面表')

    SUBFRAME_COLUMNS = [
        ('主镜头', 'master_id', 12),
        ('子画面', 'id', 14),
        ('标签', 'label', 12),
        ('关键画面', 'is_key_frame_text', 10),
        ('画面功能', 'frame_role', 14),
        ('开始(s)', 'time_start', 10),
        ('结束(s)', 'time_end', 10),
        ('进入方式', 'transition_in', 16),
        ('景别', 'shot_size', 14),
        ('机位/运镜', 'camera_movement', 24),
        ('画面构图', 'composition', 42),
        ('人物情绪', 'expression', 36),
        ('动作', 'action', 36),
        ('台词/OV/OS', 'dialogue', 44),
        ('光影', 'lighting', 30),
        ('一致性锚点', 'continuity_anchor', 34),
    ]

    for col_idx, (header, _, width) in enumerate(SUBFRAME_COLUMNS, 1):
        cell = ws_sub.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        ws_sub.column_dimensions[get_column_letter(col_idx)].width = width
    ws_sub.row_dimensions[1].height = 25
    ws_sub.freeze_panes = 'A2'

    sub_row = 2
    for shot in shots:
        for sub in shot.get('subframes', []):
            vals = {
                'master_id': shot.get('id', ''),
                'id': sub.get('id', ''),
                'label': sub.get('label', ''),
                'is_key_frame_text': '是' if sub.get('is_key_frame') else '否',
                'frame_role': sub.get('frame_role', ''),
                'time_start': sub.get('time_start', ''),
                'time_end': sub.get('time_end', ''),
                'transition_in': sub.get('transition_in', ''),
                'shot_size': sub.get('shot_size', ''),
                'camera_movement': sub.get('camera_movement', ''),
                'composition': sub.get('composition', ''),
                'expression': sub.get('expression', ''),
                'action': sub.get('action', ''),
                'dialogue': sub.get('dialogue', ''),
                'lighting': sub.get('lighting', ''),
                'continuity_anchor': sub.get('continuity_anchor', ''),
            }
            for col_idx, (_, key, _) in enumerate(SUBFRAME_COLUMNS, 1):
                cell = ws_sub.cell(row=sub_row, column=col_idx, value=vals.get(key, ''))
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border
            ws_sub.row_dimensions[sub_row].height = 60
            sub_row += 1

    if sub_row > 2:
        ws_sub.auto_filter.ref = f'A1:{get_column_letter(len(SUBFRAME_COLUMNS))}{sub_row - 1}'

    # ═══════════════════════════════════════════════════════════════
    # Sheet 4: 单镜结构化提示词
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('单镜提示词')

    # Title
    ws3.merge_cells('A1:K1')
    ws3['A1'] = f'{script_title} — 结构化单镜提示词'
    ws3['A1'].font = title_font
    ws3['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws3.merge_cells('A2:K2')
    ws3['A2'] = (
        f'画幅: {aspect_ratio} | 风格: {style_name} | 引擎: {render_engine} | '
        f'单镜时长上限: {max_duration}s | 总镜数: {len(shots)} | '
        f'子画面: {total_subframes} | 架构: 场级固定前缀 + 单镜动态参数 + 子画面关键画面 + 统一负面词 → 完整拼接'
    )
    ws3['A2'].font = meta_font
    ws3['A2'].alignment = Alignment(vertical='center')

    PROMPT_COLUMNS = [
        ('镜号', 'id', 10),
        ('时长(s)', 'duration', 7),
        ('景别', 'shot_size', 18),
        ('运镜', 'camera_movement', 28),
        ('场级固定前缀', 'scene_fixed_prefix', 45),
        ('单镜动态提示词(含分镜图)', 'shot_dynamic_prompt', 60),
        ('子画面关键画面', 'subframes_prompt', 60),
        ('统一负面词', 'unified_negatives', 35),
        ('校验状态', 'validation_status', 12),
        ('校验备注', 'validation_notes', 45),
        ('完整拼接提示词', 'full_prompt', 80),
    ]

    for col_idx, (header, _, width) in enumerate(PROMPT_COLUMNS, 1):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws3.row_dimensions[4].height = 22
    ws3.freeze_panes = 'A5'

    for col_idx, (_, _, width) in enumerate(PROMPT_COLUMNS, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, shot in enumerate(shots):
        r = row_idx + 5

        shot_scene_fixed = shot.get('scene_fixed_prefix', scene_fixed_prompt)
        shot_dynamic = shot.get('shot_dynamic_prompt', '')
        shot_subframes = shot.get('subframes_prompt', '')
        shot_negatives = shot.get('unified_negatives', unified_negatives)
        full_prompt = _compose_full_prompt(shot, scene_fixed_prompt, unified_negatives)

        vals = [
            shot.get('id', ''),
            shot.get('duration', ''),
            shot.get('shot_size', ''),
            shot.get('camera_movement', ''),
            shot_scene_fixed,
            shot_dynamic,
            shot_subframes,
            shot_negatives,
            shot.get('validation_status', ''),
            shot.get('validation_notes', ''),
            full_prompt,
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col_idx, value=v)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
        ws3.row_dimensions[r].height = 120

    # Highlight the full prompt column
    for row_idx in range(len(shots)):
        r = row_idx + 5
        ws3.cell(row=r, column=11).fill = highlight_fill

    # ═══════════════════════════════════════════════════════════════
    # Sheet 5: 完整拼接（仅完整提示词，方便复制投喂 AI）
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('完整拼接')

    ws4.merge_cells('A1:E1')
    ws4['A1'] = f'{script_title} — 完整AI提示词（仅复检通过版可直接复制投喂）'
    ws4['A1'].font = title_font
    ws4['A1'].alignment = Alignment(horizontal='left', vertical='center')

    SUB_COLUMNS = [
        ('镜号', 'id', 10),
        ('时长(s)', 'duration', 7),
        ('校验状态', 'validation_status', 12),
        ('校验备注', 'validation_notes', 36),
        ('完整提示词（复制即用）', 'full_prompt', 100),
    ]

    for col_idx, (header, _, width) in enumerate(SUB_COLUMNS, 1):
        cell = ws4.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws4.row_dimensions[3].height = 22
    ws4.freeze_panes = 'A4'

    for col_idx, (_, _, width) in enumerate(SUB_COLUMNS, 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = width

    export_row = 4
    skipped_unapproved = []
    for row_idx, shot in enumerate(shots):
        if not _is_validation_approved(shot):
            skipped_unapproved.append(shot.get('id', f'row-{row_idx + 1}'))
            continue

        r = export_row
        full_prompt = str(shot.get('full_prompt', '')).strip()

        ws4.cell(row=r, column=1, value=shot.get('id', '')).font = body_font
        ws4.cell(row=r, column=1).alignment = center_align
        ws4.cell(row=r, column=1).border = thin_border

        ws4.cell(row=r, column=2, value=shot.get('duration', '')).font = body_font
        ws4.cell(row=r, column=2).alignment = center_align
        ws4.cell(row=r, column=2).border = thin_border

        ws4.cell(row=r, column=3, value=shot.get('validation_status', '')).font = body_font
        ws4.cell(row=r, column=3).alignment = center_align
        ws4.cell(row=r, column=3).border = thin_border

        ws4.cell(row=r, column=4, value=shot.get('validation_notes', '')).font = body_font
        ws4.cell(row=r, column=4).alignment = body_align
        ws4.cell(row=r, column=4).border = thin_border

        ws4.cell(row=r, column=5, value=full_prompt).font = body_font
        ws4.cell(row=r, column=5).alignment = body_align
        ws4.cell(row=r, column=5).border = thin_border
        ws4.cell(row=r, column=5).fill = highlight_fill

        ws4.row_dimensions[r].height = 120
        export_row += 1

    if export_row == 4:
        raise ValueError(
            'No shots passed final reviewed export requirements. '
            'Sheet 5 only accepts shots with validation_status=通过, '
            'non-empty validation_notes, and a reviewed full_prompt.'
        )

    # ── Save ──
    if output_dir is None:
        raise ValueError(
            'output_dir is required: ask the user for the final export directory before generation'
        )
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    prefix = filename_prefix or script_title
    aspect_slug = '9x16' if '9:16' in aspect_ratio else '16x9'
    filename = f'{prefix}_AI提示词_{aspect_slug}.xlsx'
    output_path = output_dir / filename

    # Handle duplicate filenames
    counter = 1
    while output_path.exists():
        filename = f'{prefix}_AI提示词_{aspect_slug}_{counter}.xlsx'
        output_path = output_dir / filename
        counter += 1

    wb.save(str(output_path))
    return output_path


def build_shot_data_from_analysis(analysis: dict) -> dict:
    """
    Convert the skill's internal analysis/script data into the format
    expected by generate_xlsx().

    This is the main entry point called from the skill after all
    analysis and user confirmations are complete.

    Args:
        analysis: Dict with all data gathered during the 7-step workflow:
            - script_text: raw script content
            - meta: {script_title, aspect_ratio, style, render_engine, max_duration,
                     genre, era, conflict_type}
            - characters: [{name, age, identity, appearance, personality, voice, ...}]
            - scenes: [{name, indoor_outdoor, time_of_day, props, atmosphere}]
            - lighting: {source_direction, color_temp, contrast, fill_light, effects}
            - scene_fixed_prompt: str
            - unified_negatives: str
            - special_negative_types: {shot_type: negative_words}
            - shots: [{id, scene, duration, shot_size, camera_movement, composition,
                       expression, action, interaction, dialogue, audio, rhythm,
                       lighting_enhancement, tier, special_negatives, ...}]

    Returns:
        Dict ready for generate_xlsx()
    """
    shots = analysis.get('shots', [])
    scene_fixed_prompt = analysis.get('scene_fixed_prompt', '')
    unified_negatives = analysis.get('unified_negatives', '')

    # Enrich each shot with scene-level fixed fields
    enriched_shots = []
    for s in shots:
        enriched = _enrich_shot_with_subframes(s)
        enriched['scene_fixed_prefix'] = s.get('scene_fixed_prefix', scene_fixed_prompt)
        enriched['unified_negatives'] = s.get('unified_negatives', unified_negatives)
        enriched['subframes_prompt'] = _build_visual_subframes_prompt(enriched.get('subframes', []))
        enriched['subframes_inline'] = _build_visual_subframes_inline(enriched.get('subframes', []))

        enriched['shot_dynamic_prompt'] = _build_visual_dynamic_prompt(enriched)

        enriched['full_prompt'] = _compose_full_prompt(
            enriched,
            enriched.get('scene_fixed_prefix', scene_fixed_prompt),
            enriched.get('unified_negatives', unified_negatives),
        )

        enriched_shots.append(enriched)

    return {
        'meta': analysis.get('meta', {}),
        'characters': analysis.get('characters', []),
        'scenes': analysis.get('scenes', []),
        'lighting': analysis.get('lighting', {}),
        'scene_fixed_prompt': scene_fixed_prompt,
        'unified_negatives': unified_negatives,
        'special_negative_types': analysis.get('special_negative_types', {}),
        'shots': enriched_shots,
    }


def _build_dynamic_prompt(shot: dict) -> str:
    """Build the single-shot dynamic prompt from shot fields."""
    parts = []

    if shot.get('characters_present'):
        parts.append(f'【出场人物】{shot["characters_present"]}')

    if shot.get('organization_mode'):
        parts.append(f'【组织模式】{shot["organization_mode"]}')

    # Basic params
    basic = []
    if shot.get('shot_size'):
        basic.append(f"景别：{shot['shot_size']}")
    if shot.get('duration'):
        basic.append(f"单镜时长：{shot['duration']}秒")
    if shot.get('camera_movement'):
        basic.append(f"运镜：{shot['camera_movement']}")
    if basic:
        parts.append('【基础参数】' + '，'.join(basic))

    # Composition
    if shot.get('composition'):
        parts.append(f'【画面构图】{shot["composition"]}')

    # Expression
    if shot.get('expression'):
        parts.append(f'【人物神态】{shot["expression"]}')

    # Action/Interaction
    if shot.get('action') or shot.get('interaction'):
        action_text = shot.get('action', '')
        if shot.get('interaction'):
            action_text += f'，交互细节：{shot["interaction"]}'
        parts.append(f'【动作交互】{action_text}')

    # Dialogue/Audio
    if shot.get('dialogue') or shot.get('audio'):
        dialogue_text = shot.get('dialogue', '')
        if shot.get('audio'):
            dialogue_text += f'；{shot["audio"]}'
        parts.append(f'【台词音频】{dialogue_text}')

    # Rhythm
    if shot.get('rhythm'):
        parts.append(f'【运镜节奏】{shot["rhythm"]}')

    # Lighting enhancement
    if shot.get('lighting_enhancement'):
        parts.append(f'【光影强化】{shot["lighting_enhancement"]}')

    # Subframes / keyframes
    if shot.get('subframes_prompt'):
        parts.append(shot['subframes_prompt'])

    return '\n'.join(parts)


# ── CLI entry ──
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Split Storyboard To Script — XLSX Generator'
    )
    parser.add_argument(
        'input_json',
        help='Path to analysis JSON file (output from skill analysis phase)',
    )
    parser.add_argument(
        '--output-dir', '-o',
        required=True,
        help='User-confirmed final export directory (required; no default path)',
    )
    parser.add_argument(
        '--prefix', '-p',
        default=None,
        help='Filename prefix (default: script_title from meta)',
    )

    args = parser.parse_args()

    with open(args.input_json, 'r', encoding='utf-8-sig') as f:
        analysis = json.load(f)

    data = build_shot_data_from_analysis(analysis)
    output_dir = Path(args.output_dir)
    path = generate_xlsx(data, output_dir=output_dir, filename_prefix=args.prefix)

    print(f'[OK] XLSX written: {path}')
