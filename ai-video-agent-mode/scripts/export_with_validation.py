#!/usr/bin/env python3
"""Export a normalized current-contract package, then run the 30-point gate.

Usage:
    python3 export_with_validation.py <user_confirmed_export_md_path> <run_dir>
    python3 export_with_validation.py --regenerate <user_confirmed_export_md_path> <run_dir>

The output path is mandatory and must come from the current user's explicit
confirmation. This script never invents an export location.
"""

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(__file__))
from normalize_prompt_package import normalize_package
from materialize_master_tasks import materialize as materialize_master_tasks
from validate_master_tasks import validate as validate_master_tasks
from prompt_contract import (
    jimeng_feed_prompt,
    production_control_grounding_issues,
    production_control_grounding_report,
)
from direct_prompt_compiler import compile_direct_prompt, compile_director_card
from spatial_storyboard import build_spatial_storyboard_reference
from current_keyframe import build_keyframe_sequence
from emotion_camera_audit import audit as audit_emotion_camera
from episode_state_graph import build_episode_state_graph
from episode_director_audit import audit as audit_episode_director
from pipeline_state import AGENT_PHASES, load_state
from record_batch_provenance import verify as verify_provenance
from pipeline_runtime import atomic_json
from production_intelligence import build_sentence_provenance, predict_action_failure
from contract_registry import PROMPT_CONTRACT_VERSION
from seedance_target import TARGET_LABELS, adapt_lighting_text, adapt_visual_prefix, normalize_target, variant_paths
from validation_receipt import verify_receipt


CHECK_EXPORT = os.path.join(os.path.dirname(__file__), "check_export.py")


def export_with_validation(md_path, run_dir):
    package_path = _find_package(run_dir)
    if not package_path:
        raise SystemExit("Missing prompt package in run directory")
    _require_agent_dispatch_gates(run_dir, package_path)
    source_sha256 = _sha256(package_path)
    normalize_package(package_path, package_path)
    _record_normalization_provenance(package_path, source_sha256)
    master_path, master_package = materialize_master_tasks(run_dir, source_path=package_path)
    master_issues = validate_master_tasks(run_dir)
    if master_issues:
        raise SystemExit("Invalid main-shot delivery package: " + "; ".join(master_issues[:8]))
    receipt_ok, receipt_reason, _receipt = verify_receipt(run_dir, package_path)
    if receipt_ok:
        print("[EXPORT] verified validation receipt reused: " + receipt_reason)
    else:
        print("[EXPORT] validation receipt unavailable: %s; running full gates" % receipt_reason)
        episode_graph, episode_graph_path = build_episode_state_graph(run_dir)
        if not episode_graph.get("pass"):
            print("[EXPORT] DELIVERY BLOCKED - episode state graph failed: " + episode_graph_path)
            return 1
        director_audit, director_audit_path = audit_episode_director(run_dir)
        if not director_audit.get("pass"):
            print("[EXPORT] DELIVERY BLOCKED - episode director audit failed: " + director_audit_path)
            return 1
        emotion_audit, emotion_audit_path = audit_emotion_camera(run_dir)
        if not emotion_audit.get("pass"):
            print("[EXPORT] DELIVERY BLOCKED - emotion/camera audit failed: " + emotion_audit_path)
            return 1
        quality = subprocess.run([sys.executable, CHECK_EXPORT, "--quality", run_dir], text=True, capture_output=True)
        if quality.stdout:
            print(quality.stdout, end="")
        if quality.stderr:
            print(quality.stderr, file=sys.stderr, end="")
        if quality.returncode:
            print("[EXPORT] DELIVERY BLOCKED - quality gate failed before writing deliverables")
            return quality.returncode
    package = _load(package_path)
    plan = _load(os.path.join(run_dir, ".cache", "orchestrator", "shot_plan.json"))
    config = _load(os.path.join(run_dir, "project_config.json"))
    seedance_target = normalize_target(config.get("seedance_target", "auto"))
    feed_paths, index_path = variant_paths(md_path, seedance_target)
    destination_dir = os.path.dirname(os.path.abspath(md_path))
    temp_dir = tempfile.mkdtemp(prefix=".jimeng-export-", dir=destination_dir)
    temporary_feeds = {
        target: os.path.join(temp_dir, os.path.basename(path))
        for target, path in feed_paths.items()
    }
    temporary_xlsx = os.path.join(temp_dir, os.path.basename(os.path.splitext(md_path)[0]) + ".xlsx")
    temporary_concise = os.path.join(temp_dir, os.path.basename(os.path.splitext(md_path)[0]) + ".concise.md")
    temporary_engineering = os.path.join(temp_dir, os.path.basename(os.path.splitext(md_path)[0]) + ".engineering.md")
    compile_reports = []
    try:
        for target, temporary_md in temporary_feeds.items():
            target_reports = []
            _write_master_markdown(temporary_md, master_package, plan, target_reports, target)
            compile_reports.extend(target_reports)
        xlsx_written = _write_workbook(temporary_xlsx, package, plan, {})
        if xlsx_written:
            for temporary_md in temporary_feeds.values():
                validation_xlsx = os.path.splitext(temporary_md)[0] + ".xlsx"
                if os.path.abspath(validation_xlsx) != os.path.abspath(temporary_xlsx):
                    shutil.copyfile(temporary_xlsx, validation_xlsx)
        for target, temporary_md in temporary_feeds.items():
            result = subprocess.run(
                [sys.executable, CHECK_EXPORT, temporary_md, run_dir], text=True, capture_output=True,
            )
            if result.stdout:
                print(result.stdout, end="")
            if result.stderr:
                print(result.stderr, file=sys.stderr, end="")
            if result.returncode:
                print("[EXPORT] DELIVERY BLOCKED - temporary deliverables discarded")
                return result.returncode
        if seedance_target != "both":
            only_target = next(iter(feed_paths))
            _write_concise_markdown(temporary_concise, master_package, plan, only_target)
            _write_engineering_review(temporary_engineering, master_package, plan, compile_reports)
        temporary_index = ""
        if seedance_target == "both":
            temporary_index = os.path.join(temp_dir, os.path.basename(index_path))
            _write_target_index(temporary_index, plan, feed_paths)
        for target, destination in feed_paths.items():
            os.replace(temporary_feeds[target], destination)
        if seedance_target == "both":
            os.replace(temporary_index, index_path)
        else:
            concise_path = os.path.splitext(md_path)[0] + ".concise.md"
            engineering_path = os.path.splitext(md_path)[0] + ".engineering.md"
            os.replace(temporary_concise, concise_path)
            os.replace(temporary_engineering, engineering_path)
        if xlsx_written:
            os.replace(temporary_xlsx, os.path.splitext(md_path)[0] + ".xlsx")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
    compile_report_path = os.path.join(run_dir, ".cache", "export", "direct_prompt_compile_report.json")
    atomic_json(compile_report_path, {
        "contract_version": PROMPT_CONTRACT_VERSION,
        "pass": True,
        "shot_count": len(compile_reports),
        "shots": compile_reports,
    })
    _record_export_result(run_dir, md_path, compile_report_path, seedance_target, feed_paths, index_path)
    print("[EXPORT] DELIVERY APPROVED")
    for target, path in feed_paths.items():
        print("[EXPORT] Markdown %s: %s" % (TARGET_LABELS[target], path))
    if index_path:
        print("[EXPORT] Non-feed index: " + index_path)
    else:
        print("[EXPORT] Concise director cards: " + os.path.splitext(md_path)[0] + ".concise.md")
        print("[EXPORT] Engineering review: " + os.path.splitext(md_path)[0] + ".engineering.md")
    print("[EXPORT] Master tasks: " + master_path)
    if xlsx_written:
        print("[EXPORT] XLSX: " + os.path.splitext(md_path)[0] + ".xlsx")
    else:
        print("[EXPORT] XLSX skipped: openpyxl is unavailable; Markdown delivery is complete")
    return 0


def _record_export_result(run_dir, md_path, compile_report_path="", seedance_target="auto", feed_paths=None, index_path=""):
    package_path = _find_package(run_dir)
    if not package_path:
        return
    destination = os.path.abspath(md_path)
    feed_paths = {key: os.path.abspath(value) for key, value in (feed_paths or {seedance_target: destination}).items()}
    primary_path = next(iter(feed_paths.values())) if len(feed_paths) == 1 else os.path.abspath(index_path)
    atomic_json(os.path.join(run_dir, ".cache", "export", "result.json"), {
        "pass": True,
        "exported_at": time.time(),
        "seedance_target": seedance_target,
        "markdown_path": primary_path,
        "markdown_sha256": _sha256(primary_path),
        "markdown_paths": feed_paths,
        "markdown_sha256_by_target": {key: _sha256(value) for key, value in feed_paths.items()},
        "index_markdown_path": os.path.abspath(index_path) if index_path else "",
        "package_sha256": _sha256(package_path),
        "xlsx_path": os.path.splitext(destination)[0] + ".xlsx",
        "concise_markdown_path": os.path.splitext(destination)[0] + ".concise.md" if seedance_target != "both" else "",
        "engineering_markdown_path": os.path.splitext(destination)[0] + ".engineering.md" if seedance_target != "both" else "",
        "direct_prompt_compile_report": os.path.abspath(compile_report_path) if compile_report_path else "",
    })


def _require_agent_dispatch_gates(run_dir, package_path):
    """Refuse delivery unless the current package descends from verified workers."""
    state = load_state(run_dir)
    incomplete = [
        phase for phase in AGENT_PHASES
        if state.get("phases", {}).get(phase, {}).get("status") != "done"
    ]
    if incomplete:
        raise SystemExit("DISPATCH_GATE: Agent phases are incomplete: " + ", ".join(sorted(incomplete)))
    manifest_path = package_path + ".merge_provenance.json"
    manifest = _load_optional(manifest_path)
    if not manifest or manifest.get("output_path") != os.path.abspath(package_path):
        raise SystemExit("DISPATCH_GATE: verified merge provenance is required before export")
    if manifest.get("output_sha256") != _sha256(package_path):
        raise SystemExit("DISPATCH_GATE: prompt package changed after verified merge")
    sources = manifest.get("source_batches")
    if not isinstance(sources, list) or not sources:
        raise SystemExit("DISPATCH_GATE: merge provenance has no verified worker batches")
    for source in sources:
        batch_path = source.get("batch_path") if isinstance(source, dict) else ""
        valid, reason, _record = verify_provenance(batch_path) if batch_path and os.path.exists(batch_path) else (False, "batch missing", None)
        if not valid:
            raise SystemExit("DISPATCH_GATE: source worker batch is invalid: " + reason)


def _sha256(path):
    import hashlib
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _record_normalization_provenance(package_path, source_sha256):
    """Record the deterministic negative-prompt injection in the merge ledger."""
    manifest_path = package_path + ".merge_provenance.json"
    manifest = _load_optional(manifest_path)
    if not manifest:
        raise SystemExit("DISPATCH_GATE: normalized package requires merge provenance")
    if manifest.get("output_sha256") != source_sha256:
        raise SystemExit("DISPATCH_GATE: package changed before deterministic normalization")
    current_hash = _sha256(package_path)
    manifest["output_sha256"] = current_hash
    manifest["normalization"] = {
        "name": "normalize_prompt_package",
        "input_sha256": source_sha256,
        "output_sha256": current_hash,
        "recorded_at": time.time(),
    }
    atomic_json(manifest_path, manifest)


def _write_master_markdown(path, master_package, plan, compile_reports=None, seedance_target="auto"):
    """Write the user-facing delivery from canonical main-shot tasks."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    masters = {item.get("shot_id", ""): item for item in master_package.get("shots", []) if isinstance(item, dict)}
    lines = [
        f"# {plan.get('project_name', '')} 即梦投喂分镜｜{TARGET_LABELS[seedance_target]}", "",
        "## 使用说明", "",
        "- `【画面描述｜直接复制】` 是每个主镜头推荐复制到即梦正文框的正向提示词。", "",
        "- `【负面提示词｜直接复制】` 可直接复制到负面词框；高风险镜头会额外给出本镜必要约束和补充负面词。", "",
        "- `【画面参数】`、`【运镜描述】`、`【光影描述】` 用于人工复核画幅、影调、色卡、镜头路径和光影质感，不是额外提示词段落。", "",
        f"- Seedance 目标：`{seedance_target}`；本文件可独立投喂，不与另一版本拼接。", "",
        "## 全局锁定", "",
        f"- 全局风格：{plan.get('canvas', '')}画幅，{plan.get('visual_style', '')}，即梦 T2V。", "",
    ]
    lines.extend(_global_lock_lines(master_package, plan))
    lines.extend([
        "## 制作质量总控", "",
    ])
    lines.extend(_global_quality_control_lines(master_package))
    lines.extend([
        "## 通用负面提示词｜直接复制", "",
        _global_negative_prompt(master_package), "",
        "## 场景状态表", "",
    ])
    lines.extend(_scene_state_lines(master_package, plan))
    lines.extend(["## 分镜投喂卡", ""])
    current_scene = None
    planned_shots = plan.get("shots", []) if isinstance(plan.get("shots", []), list) else []
    for index, planned in enumerate(planned_shots):
        scene = planned.get("scene", "场景")
        if scene != current_scene:
            lines.extend([f"**{scene}**", ""])
            current_scene = scene
        task = masters.get(planned.get("shot_id", ""))
        if not task:
            continue
        next_task = None
        for following in planned_shots[index + 1:]:
            candidate = masters.get(following.get("shot_id", ""))
            if candidate:
                next_task = candidate
                break
        control = task.get("generation_control", {}) if isinstance(task.get("generation_control"), dict) else {}
        metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
        camera_beats = metadata.get("camera_beat_map", []) if isinstance(metadata.get("camera_beat_map"), list) else []
        lines.extend([
            f"#### {task.get('shot_id', '')}｜镜头组总时长：{float(task.get('duration', 0) or 0):g}s", "",
            "【出现人物】", "",
            _visible_people(metadata), "",
            "【镜号】", "",
            f"1，{float(task.get('duration', 0) or 0):g}s，{'复杂' if _high_risk_direct_blocks_enabled(task) else '普通'}。", "",
            "【画面参数】", "",
            _picture_parameter_line(task, plan), "",
            "【画面描述｜直接复制】", "",
            _build_direct_copy_prompt(task, plan, compile_reports, seedance_target), "",
            "【导演卡｜直接复制｜≤500字】", "",
            _build_director_card(task, plan, compile_reports, seedance_target), "",
            "【运镜描述】", "",
            _camera_description(camera_beats, task), "",
            "【光影描述】", "",
            adapt_lighting_text(_lighting_description(task), seedance_target), "",
            "【负面提示词｜直接复制】", "",
            str(task.get("negative_prompt", "") or ""), "",
            "【表演与声音】", "",
        ])
        events = metadata.get("dialogue_events", []) if isinstance(metadata.get("dialogue_events"), list) else []
        if events:
            lines.extend([
                "| 引用 | 类型 | 人物 | 逐字原文 | 时间窗 | 台词功能 | 潜台词 | 原文重音词 | 潜台词可见证据 | 轮次关系 | 会话模式 | 响应延迟 | 抢话/打断窗口 | 会话源文依据 | 神态 | 身体状态 | 语气 | 气口 | 口型同步 |",
                "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
            ])
            for event in events:
                lines.append("| " + " | ".join(_dialogue_md_cell(event, field) for field in (
                    "ref", "kind", "speaker", "text", "time_range", "line_function", "subtext",
                    "stress_words", "subtext_visible_evidence", "turn_relation", "conversation_mode", "response_latency",
                    "overlap_or_interrupt_window", "conversation_source_basis", "facial_state", "body_state",
                    "delivery", "breath_pause_plan", "lip_sync",
                )) + " |")
            lines.append("")
        else:
            lines.extend(["无台词。", ""])
        lines.extend(["【状态继承】", "", _state_carryover(task), ""])
        lines.extend(["【剪辑衔接】", "", _build_transition_prompt(task, next_task), ""])
        lines.extend(["【本镜制作控制】", "", _shot_production_control(task), ""])
        if camera_beats:
            lines.extend(["【镜头执行】", ""])
            _append_execution_beats(lines, {"camera_beat_map": camera_beats})
        if _high_risk_direct_blocks_enabled(task):
            lines.extend([
                "【本镜必要约束｜直接复制】", "", _build_direct_constraint_block(task), "",
                "【本镜补充负面提示词｜直接复制】", "", _build_direct_negative_block(task), "",
            ])
        lines.extend(["【内部溯源】", "", "来源子镜：" + "、".join(task.get("source_subshot_ids", [])), ""])
        keyframe_sequence = build_keyframe_sequence(task, planned, plan.get("canvas", "16:9"), plan.get("visual_style", ""))
        if keyframe_sequence:
            lines.extend([
                "【关键帧生图提示】", "",
                f"优先级：{keyframe_sequence['priority']}；触发原因：{keyframe_sequence['reason']}。", "",
            ])
            for frame in keyframe_sequence["frames"]:
                lines.extend([f"{frame['label']}｜{frame['time_seconds']:g}s", "", frame["prompt"], ""])
            lines.extend([
                "【即梦视频提示｜配合关键帧】", "",
                keyframe_sequence["video_prompt"], "",
                "【人物/道具状态差异表】", "",
                "| 对象 | 起始状态 | 戏眼状态 | 结束状态 |",
                "|---|---|---|---|",
            ])
            for row in keyframe_sequence["state_diff"]:
                lines.append("| %s | %s | %s | %s |" % tuple(
                    _md_cell(row.get(field, "")) for field in ("subject", "start", "dramatic", "end")
                ))
            lines.extend(["", "【关键帧连续性与T2V事实校验】", ""])
            for check in keyframe_sequence["continuity_check"] + keyframe_sequence["fact_consistency"]:
                lines.append("- %s：%s（%s）" % (
                    check["name"], "通过" if check["pass"] else "失败", check["evidence"]
                ))
            lines.extend([
                "", "【关键帧负面提示词｜直接复制】", "",
                keyframe_sequence["negative_prompt"], "",
            ])
        spatial_reference = build_spatial_storyboard_reference(task, planned, plan.get("canvas", "16:9"), plan.get("visual_style", ""))
        if spatial_reference:
            lines.extend([
                f"【空间与道具锁定｜{spatial_reference['priority']}生成】", "",
                f"触发原因：{spatial_reference['reason']}。", "",
                "俯视空间调度图：" + spatial_reference["overhead_map_prompt"], "",
                "人物站位姿态参考图：" + spatial_reference["blocking_board_prompt"], "",
                "分镜图负面提示词：" + spatial_reference["negative_prompt"], "",
            ])
        lines.extend(["---", ""])
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _write_concise_markdown(path, master_package, plan, seedance_target="auto"):
    """Write only compact copy cards and basic shot identity."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    masters = {item.get("shot_id", ""): item for item in master_package.get("shots", []) if isinstance(item, dict)}
    lines = [
        "# %s 即梦导演卡（简洁交付视图）" % plan.get("project_name", ""), "",
        "> 仅含可复制导演卡；完整合同、QA和溯源见同名 engineering.md / xlsx。", "",
    ]
    current_scene = None
    for planned in plan.get("shots", []):
        scene = planned.get("scene", "场景")
        if scene != current_scene:
            lines.extend(["## " + str(scene), ""])
            current_scene = scene
        task = masters.get(planned.get("shot_id", ""))
        if not task:
            continue
        card = _build_director_card(task, plan, seedance_target=seedance_target)
        lines.extend([
            "### %s｜%ss" % (task.get("shot_id", ""), float(task.get("duration", 0) or 0)), "",
            "【导演卡｜直接复制｜≤500字】", "", card, "",
            "【负面提示词｜直接复制】", "", str(task.get("negative_prompt", "") or ""), "", "---", "",
        ])
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _write_target_index(path, plan, feed_paths):
    """Write a human-facing comparison entry point; never use it as a feed prompt."""
    speech_contract = "台词/OS/OV/系统音" if _plan_has_system_sound(plan) else "台词/OS/OV"
    lines = [
        "# 双版本索引｜%s" % plan.get("project_name", "即梦分镜"), "",
        "> 本文件只用于选择版本和核对合同，不要复制到即梦。两份分镜来自同一份镜头、台词、人物、空间、道具与时长合同。", "",
        "## 可投喂文件", "",
        "| 目标 | 文件 | 光影适配 | 使用场景 |", "|---|---|---|---|",
        "| Seedance 2.0 | `%s` | 简洁动机光，浅至中等阴影，面部可读性优先 | 稳定兼容、对白和低风险动作 |" % os.path.basename(feed_paths["2.0"]),
        "| Seedance 2.5 | `%s` | 更完整的动机光、明暗层次、局部环境色与高光滚降 | 需要更强光影表现时 |" % os.path.basename(feed_paths["2.5"]),
        "", "## 不变合同", "",
        "- 镜号、镜头顺序、时长、%s 原文、人物、空间锚点、道具归属和剧情因果保持一致。" % speech_contract,
        "- 两份文件都可独立投喂；不要把两份正文拼接后投喂。",
        "- `seedance_target=both` 的触发点是项目配置确认时选择 `both`，随后照常完成一次管线运行并在导出阶段自动生成本表和两份版本。",
        "",
    ]
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _write_engineering_review(path, package, plan, compile_reports):
    """Write an audit view with sentence provenance and repair-oriented metrics."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    lines = [
        "# %s 工程审查视图" % plan.get("project_name", ""), "",
        "> 此文件只供 QA/审计，不应投喂即梦。包含句级来源、动作失败预测、压缩省略和合同字段。", "",
    ]
    by_id = {item.get("subshot_id", ""): item for item in compile_reports if isinstance(item, dict)}
    for shot in package.get("shots", []):
        if not isinstance(shot, dict):
            continue
        sid = shot.get("subshot_id", "")
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        report = by_id.get(sid, {})
        lines.extend([
            "## %s" % sid, "",
            "- 完整直投字数：%s" % report.get("char_count", "未记录"),
            "- 导演卡字数：%s" % report.get("director_card_char_count", "未记录"),
            "- 动作失败预测：%s（%s）" % (
                (report.get("action_failure_prediction", {}) or {}).get("risk_level", "未记录"),
                "、".join((report.get("action_failure_prediction", {}) or {}).get("reasons", [])),
            ),
            "- 句级来源：" + json.dumps(report.get("sentence_provenance", []), ensure_ascii=False),
            "- 压缩省略：" + json.dumps(report.get("director_card_omitted", []), ensure_ascii=False),
            "- 活动合同：" + ", ".join(key for key, value in metadata.items() if isinstance(value, dict) and value),
            "- 审美优先级：" + json.dumps(metadata.get("aesthetic_priority", {}), ensure_ascii=False),
            "- 静态美学合同：" + json.dumps(metadata.get("static_aesthetic_contract", {}), ensure_ascii=False),
            "- 动态美学合同：" + json.dumps(metadata.get("dynamic_aesthetic_contract", {}), ensure_ascii=False),
            "- 真实候选审美评分：未执行；仅在提供真实图片/视频并明确要求复核时记录。",
            "- 审美复核清单：构图焦点、动机光与曝光、色彩分离、材质层次、记忆帧；动态另查起始/变化/终态、运动动机、缓急、稳定性和视觉落点。",
            "",
        ])
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _write_markdown(path, package, plan, director):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    shots = package.get("shots", [])
    by_id = {shot.get("subshot_id", ""): shot for shot in shots}
    director_map = {
        item.get("subshot_id", ""): item
        for item in director.get("items", []) if isinstance(item, dict) and item.get("subshot_id")
    }
    ordered_ids = [
        subshot.get("subshot_id", "")
        for plan_shot in plan.get("shots", [])
        for subshot in plan_shot.get("subshots", [])
        if subshot.get("subshot_id", "") in by_id
    ]
    lines = [
        f"# {plan.get('project_name', '')} 即梦 T2V 提示词包",
        "",
        f"画幅：{plan.get('canvas', '')} | 风格：{plan.get('visual_style', '')}",
        "",
        "> 每个主镜头是一项即梦 T2V 生成任务；正文含至多 3 个可见子镜。内部 QA 数据保留在 XLSX。",
        "",
        "---",
        "",
    ]
    current_scene = None
    ordered_index = {sid: idx for idx, sid in enumerate(ordered_ids)}
    for plan_shot in plan.get("shots", []):
        scene = plan_shot.get("scene", "场景")
        if scene != current_scene:
            lines.extend([f"## {scene}", ""])
            current_scene = scene
        children = [by_id.get(subshot.get("subshot_id", "")) for subshot in plan_shot.get("subshots", [])]
        children = [child for child in children if child]
        total = sum(float(child.get("duration", 0) or 0) for child in children)
        lines.extend([f"### {plan_shot.get('shot_id', '')}（{total:g}秒）", ""])
        if children:
            control = children[0].get("generation_control", {}) if isinstance(children[0].get("generation_control"), dict) else {}
            lines.extend([
                "**即梦操作卡**",
                "",
                f"模式：即梦 T2V｜画幅：{plan.get('canvas', '')}｜时长：{total:g}秒｜原生音频：{'开启' if control.get('audio_enabled') else '关闭'}",
                "",
                "**主镜头即梦提示词**",
                "",
                "```text",
                _build_master_prompt(children, plan),
                "```",
                "",
                "**主镜头负面提示词**",
                "",
                "```text",
                _merge_negative_prompts(children),
                "```",
                "",
            ])
        for shot in children:
            metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
            dialogue_events = metadata.get("dialogue_events", []) if isinstance(metadata.get("dialogue_events"), list) else []
            next_shot = None
            idx = ordered_index.get(shot.get("subshot_id", ""), -1)
            if idx >= 0 and idx + 1 < len(ordered_ids):
                next_shot = by_id.get(ordered_ids[idx + 1])
            lines.extend([
                f"#### 子镜头 {shot.get('subshot_id', '')}｜{float(shot.get('duration', 0) or 0):g}秒",
                "",
                "**子镜执行节拍**",
                "",
            ])
            _append_execution_beats(lines, director_map.get(shot.get("subshot_id", ""), {}))
            lines.extend(["**承接下一镜**", "", _build_transition_prompt(shot, next_shot), ""])
            speech_heading = "**台词/OS/OV/系统音表演**" if any(
                event.get("kind") == "系统音" for event in dialogue_events if isinstance(event, dict)
            ) else "**台词/OS/OV表演**"
            lines.extend([
                speech_heading,
                "",
            ])
            if dialogue_events:
                lines.extend([
                    "| 引用 | 类型 | 人物 | 逐字原文 | 时间窗 | 台词功能 | 潜台词 | 原文重音词 | 潜台词可见证据 | 轮次关系 | 会话模式 | 响应延迟 | 抢话/打断窗口 | 会话源文依据 | 神态 | 身体状态 | 语气 | 气口 | 口型同步 |",
                    "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
                ])
                for event in dialogue_events:
                    lines.append("| " + " | ".join(_dialogue_md_cell(event, field) for field in (
                        "ref", "kind", "speaker", "text", "time_range", "line_function", "subtext",
                        "stress_words", "subtext_visible_evidence", "turn_relation", "conversation_mode", "response_latency",
                        "overlap_or_interrupt_window", "conversation_source_basis", "facial_state", "body_state",
                        "delivery", "breath_pause_plan", "lip_sync",
                    )) + " |")
                lines.append("")
            else:
                lines.extend(["无。", ""])
        lines.extend(["---", ""])
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _plan_has_system_sound(plan):
    events = plan.get("dialogue_events", {}) if isinstance(plan, dict) else {}
    if isinstance(events, dict):
        events = events.values()
    elif not isinstance(events, list):
        events = []
    return any(isinstance(event, dict) and event.get("kind") == "系统音" for event in events)


def _build_direct_copy_prompt(task, plan, compile_reports=None, seedance_target="auto"):
    """Build a compact Jimeng copy block from the canonical five-section prompt."""
    segments, required_fragments, information_budget = _direct_prompt_inputs(task, plan, seedance_target)
    compiled = compile_direct_prompt(
        segments, required_fragments, max_chars=700, information_budget=information_budget
    )
    if compiled["issues"]:
        raise ValueError("；".join(compiled["issues"]))
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    grounding_report = production_control_grounding_report(metadata, compiled["text"])
    grounding_issues = production_control_grounding_issues(metadata, compiled["text"])
    if grounding_issues:
        raise ValueError("；".join(grounding_issues))
    if isinstance(compile_reports, list):
        compile_reports.append({
            "shot_id": str(task.get("shot_id", "") or ""),
            "subshot_id": str(task.get("subshot_id", "") or ""),
            "char_count": len(compiled["text"]),
            "segment_order": compiled["segment_order"],
            "removed_duplicate_count": compiled["removed_duplicate_count"],
            "omitted": compiled["omitted"],
            "protected_dialogue_fragments": _dialogue_required_fragments(task),
            "protected_required_fragments": compiled["required_fragments"],
            "budget_profile": compiled["budget_profile"],
            "visual_enhancer_limit": compiled["visual_enhancer_limit"],
            "active_visual_enhancers": compiled["active_visual_enhancers"],
            "sentence_provenance": build_sentence_provenance(task, compiled["text"]),
            "action_failure_prediction": predict_action_failure(task),
            "production_control_grounding": grounding_report,
        })
    return _clean_export_direct_text(compiled["text"])


def _build_director_card(task, plan, compile_reports=None, seedance_target="auto"):
    segments, required_fragments, information_budget = _direct_prompt_inputs(task, plan, seedance_target)
    # The compact card keeps dialogue, hard facts, and the dynamic action rail;
    # the full direct-copy block is the canonical place for the static material anchor.
    card_required = _dialogue_required_fragments(task) + _budget_render_units(information_budget)
    motion_anchor = _dynamic_motion_anchor(
        task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    )
    if motion_anchor:
        card_required.append(motion_anchor)
    authored_card = str(task.get("director_card", "") or "").strip()
    if authored_card:
        # Semantic facts may be rephrased by the model in the companion card;
        # only verbatim dialogue remains an exact engineering lock.
        card_required = _dialogue_required_fragments(task)
    card_segments = (
        [{"kind": "model_authored_director_card", "text": authored_card}]
        if authored_card else segments
    )
    compiled = compile_director_card(
        card_segments, _unique_nonempty(card_required), information_budget
    )
    if compiled["issues"]:
        raise ValueError("导演卡编译失败：" + "；".join(compiled["issues"]))
    if isinstance(compile_reports, list):
        for report in reversed(compile_reports):
            if report.get("shot_id") == task.get("shot_id"):
                report["director_card_char_count"] = len(compiled["text"])
                report["director_card_segment_order"] = compiled["segment_order"]
                report["director_card_omitted"] = compiled["omitted"]
                break
    return _clean_export_direct_text(compiled["text"])


def _direct_prompt_inputs(task, plan, seedance_target="auto"):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
    budget = metadata.get("prompt_information_budget", {}) if isinstance(metadata.get("prompt_information_budget"), dict) else {}
    motion_anchor = _dynamic_motion_anchor(metadata)
    static_anchor = _static_visual_anchor(metadata)
    enhancer_limit = budget.get("visual_enhancer_limit")
    cinematic = _cinematic_direct_clause(metadata)
    video_texture = _video_texture_direct_clause(metadata, static_anchor=static_anchor if enhancer_limit == 1 and motion_anchor else "")
    if enhancer_limit == 1:
        if motion_anchor:
            cinematic = ""
        else:
            video_texture = ""
    prefix = adapt_visual_prefix(_compact(palette.get("visual_scene_prefix", "")), seedance_target)
    style_parts = [
        "%s画幅" % plan.get("canvas", "") if plan.get("canvas", "") else "",
        str(plan.get("visual_style", "") or "").strip("，。 "),
        prefix.strip("，。 "),
    ]
    style_prefix = _compact("，".join(part for part in style_parts if part) + "。")
    sections = _prompt_sections(task.get("full_prompt", ""))
    composition_clause = _composition_direct_clause(metadata)
    terminal_clause = _terminal_frame_direct_clause(metadata)
    if sections:
        space_text = sections.get("主体与空间锁定", "")
        continuity_text = sections.get("主镜头连续规则", "")
        if composition_clause and composition_clause not in space_text:
            space_text = "；".join(part for part in (space_text, composition_clause) if part)
        if terminal_clause and terminal_clause not in continuity_text:
            continuity_text = "；".join(part for part in (continuity_text, terminal_clause) if part)
        segments = [
            {"kind": "visual_prefix", "text": style_prefix},
            {"kind": "space", "text": space_text},
            {"kind": "continuity", "text": continuity_text},
            {"kind": "performance", "text": _strip_inner_shot_headings(sections.get("子镜头组", ""))},
            {"kind": "light", "text": adapt_lighting_text(sections.get("光照、声音与稳定约束", ""), seedance_target)},
            {"kind": "video_texture", "text": video_texture},
            {"kind": "cinematic", "text": cinematic},
        ]
    else:
        direct_fallback = jimeng_feed_prompt(task.get("full_prompt", ""))
        if composition_clause and composition_clause not in direct_fallback:
            direct_fallback = "；".join(part for part in (direct_fallback, composition_clause) if part)
        if terminal_clause and terminal_clause not in direct_fallback:
            direct_fallback = "；".join(part for part in (direct_fallback, terminal_clause) if part)
        segments = [
            {"kind": "visual_prefix", "text": style_prefix},
            {"kind": "performance", "text": direct_fallback},
            {"kind": "light", "text": adapt_lighting_text("", seedance_target) if seedance_target != "auto" else ""},
            {"kind": "video_texture", "text": video_texture},
            {"kind": "cinematic", "text": cinematic},
        ]
    control = task.get("generation_control", {}) if isinstance(task.get("generation_control"), dict) else {}
    required = _dialogue_required_fragments(task)
    required.extend(_budget_render_units(budget))
    # Static light/material anchors are already carried by the light section
    # and validated through production-control grounding. Requiring the full
    # combined sentence here prevents clause-level compression and duplicates
    # the same facts in video_texture. Keep causal motion, terminal state, and
    # explicit must_render units hard-protected.
    required.extend(value for value in (motion_anchor, terminal_clause) if value)
    return segments, _unique_nonempty(required), budget


def _composition_direct_clause(metadata):
    """Return one concrete composition skeleton for the model-facing prompt."""
    metadata = metadata if isinstance(metadata, dict) else {}
    static = metadata.get("static_aesthetic_contract", {}) if isinstance(metadata.get("static_aesthetic_contract"), dict) else {}
    cinematic = metadata.get("cinematic_image_contract", {}) if isinstance(metadata.get("cinematic_image_contract"), dict) else {}
    bible = metadata.get("visual_bible", {}) if isinstance(metadata.get("visual_bible"), dict) else {}
    values = [
        static.get("composition_hierarchy", ""),
        cinematic.get("composition_anchor", ""),
        bible.get("composition_grammar", ""),
    ]
    value = next((_compact(item) for item in values if _compact(item)), "")
    if not value:
        return ""
    return "构图骨架：" + _shorten(value, 110)


def _terminal_frame_direct_clause(metadata):
    """Translate terminal_frame_contract into positive, visible anti-duplication facts."""
    metadata = metadata if isinstance(metadata, dict) else {}
    contract = metadata.get("terminal_frame_contract", {})
    if not isinstance(contract, dict) or not contract:
        return ""
    def value(key):
        return _compact(contract.get(key, ""))
    pieces = []
    if value("visible_count") or value("final_slot_map"):
        pieces.append("最后20%只保留" + "、".join(part for part in (value("visible_count"), value("final_slot_map")) if part))
    for key in ("identity_visibility", "face_and_limb_separation", "prop_and_garment_state", "support_and_contact"):
        if value(key):
            pieces.append(value(key))
    if value("camera_lock"):
        pieces.append(value("camera_lock"))
    if value("light_exposure_lock"):
        pieces.append(value("light_exposure_lock"))
    if value("no_new_entrant"):
        pieces.append(value("no_new_entrant"))
    if value("no_duplicate_subject"):
        pieces.append(value("no_duplicate_subject"))
    if value("final_hold"):
        pieces.append(value("final_hold"))
    return "终端画面：" + "，".join(pieces) if pieces else ""


def _global_lock_lines(master_package, plan):
    shots = [shot for shot in master_package.get("shots", []) if isinstance(shot, dict)]
    scene_lines = []
    palette_by_space = {}
    state_lines = []
    seen_scene = set()
    for shot in shots:
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
        continuity = metadata.get("continuity_contract", {}) if isinstance(metadata.get("continuity_contract"), dict) else {}
        space_id = _compact(palette.get("space_id", "")) or _compact(shot.get("shot_id", ""))
        if space_id and space_id not in seen_scene:
            seen_scene.add(space_id)
            scene_lines.append(
                "- %s：%s；%s；%s" % (
                    space_id,
                    _compact(palette.get("space_master_sentence", "")) or "空间主锁定以本镜画面描述为准",
                    _compact(palette.get("light_texture_purpose", "")) or "光影服务本镜剧情任务",
                    _compact(continuity.get("prop_state", "")) or "无跨镜活动道具",
                )
            )
        tone = _compact(palette.get("tone_palette", ""))
        prefix = _compact(palette.get("visual_scene_prefix", ""))
        palette_text = "；".join(part for part in (tone, prefix) if part)
        palette_space = space_id or "场景"
        if palette_text:
            current = palette_by_space.get(palette_space, "")
            if not current or len(palette_text) > len(current):
                palette_by_space[palette_space] = palette_text
        carryover = _compact(continuity.get("next_carryover") or continuity.get("end_anchor", ""))
        if carryover:
            state_lines.append("- %s：%s" % (shot.get("shot_id", ""), carryover))
    lines = [
        "- 全局比例与支撑锁定：全程角色骨骼与头身比例恒定，人物真实身高和体型尺寸固定；四肢长度与关节比例稳定，地平线及消失关系稳定；人物身体主支撑点持续贴合当前承载面并保留接触阴影，站立时双脚接地，行走时步态交替接地，坐卧时臀背或躯干贴合承载面，腾空时保持起跳、空中与落地轨迹连续；同镜两人身高差、骨架和相对尺寸全程一致，人物画面投影只随物理距离连续变化，固定距离下画面占比保持稳定。",
        "",
        "- 本集空间锁定索引：", "",
    ]
    lines.extend(scene_lines or ["- 暂无独立空间索引；以各镜 `【画面参数】` 与 `【状态继承】` 为准。"])
    lines.extend(["", "- 本集影调色卡索引：", ""])
    palette_lines = [f"- {space_id}：{text}" for space_id, text in palette_by_space.items()]
    lines.extend(palette_lines or [f"- 全局：{plan.get('visual_style', '')}，色卡以各镜 Scene Lock 为准。"])
    if state_lines:
        lines.extend(["", "- 状态锁定：", ""])
        lines.extend(state_lines[:8])
    lines.append("")
    return lines


def _contract_text(metadata, key, fields, limit=300):
    contract = metadata.get(key, {}) if isinstance(metadata, dict) else {}
    contract = contract if isinstance(contract, dict) else {}
    values = []
    for field in fields:
        value = contract.get(field, "")
        if isinstance(value, list):
            value = "；".join(str(item) for item in value if str(item).strip())
        value = _compact(value)
        if value and value.lower() not in {"none", "n/a", "na"} and value not in values:
            values.append(value)
    return _shorten("；".join(values), limit) if values else ""


def _first_contract_text(shots, key, fields, fallback):
    for shot in shots:
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        value = _contract_text(metadata, key, fields, 360)
        if value:
            return value
    return fallback


def _global_quality_control_lines(master_package):
    shots = [shot for shot in master_package.get("shots", []) if isinstance(shot, dict)]
    visual = _first_contract_text(
        shots, "visual_bible",
        ("visual_thesis", "composition_grammar", "material_world", "imperfection_policy"),
        "每镜保持一个第一视觉落点、一个构图骨架和一至两项有剧情作用的材质或自然不完美。",
    )
    light = _first_contract_text(
        shots, "visual_bible",
        ("light_motivation", "contrast_exposure", "atmosphere_rule", "continuity_lock"),
        "现实或剧情动机光固定来源、方向、受光面和阴影结果；脸与活动手优先，高光不过曝、黑位可读。",
    )
    dynamic = _first_contract_text(
        shots, "dynamic_aesthetic_contract",
        ("motion_thesis", "camera_path", "tempo_easing", "stability_fallback"),
        "每镜只保留一个主体动作、一条摄影机路径和一个因果响应；按起幅、触发、响应、稳定落幅组织。",
    )
    performance = _first_contract_text(
        shots, "source_constraint_basemap",
        ("performance_baseline_lock", "emotion_micro_chain", "emotion_residue_contract"),
        "人物表演按触发、对外控制、最小可见泄露、对手反应和情绪余波组织，口型优先于运镜。",
    )
    cut_modes = []
    risk_rank = {"low": 1, "medium": 2, "high": 3}
    highest_risk, risk_reasons = "low", []
    for shot in shots:
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        cut = metadata.get("cut_decision_contract", {}) if isinstance(metadata.get("cut_decision_contract"), dict) else {}
        mode = _compact(cut.get("cut_mode", ""))
        if mode and mode not in cut_modes:
            cut_modes.append(mode)
        reroll = metadata.get("reroll_control", {}) if isinstance(metadata.get("reroll_control"), dict) else {}
        level = str(reroll.get("risk_level", "low") or "low").lower()
        if risk_rank.get(level, 0) > risk_rank.get(highest_risk, 0):
            highest_risk = level
        reason = _compact(reroll.get("risk_reason", ""))
        if reason and reason not in risk_reasons:
            risk_reasons.append(reason)
    montage = "本项目切点模式：%s；每次切换必须有动作、反应、台词、声音、匹配形态或场景变化触发，并产生新信息。" % (
        "、".join(cut_modes[:6]) if cut_modes else "保持或动作/反应切点"
    )
    risk = "全片最高抽卡风险：%s；锁定身份、人体比例、接触/支撑、空间、道具、功能面、口型、光源和稳定终态%s。" % (
        highest_risk,
        "；主要来源：" + "；".join(risk_reasons[:3]) if risk_reasons else "",
    )
    return [
        "- 画面质感基线：" + visual, "",
        "- 光效与曝光连续：" + light, "",
        "- 动态美学基线：" + dynamic, "",
        "- 表演与情绪基线：" + performance, "",
        "- 蒙太奇与剪辑基线：" + montage, "",
        "- 穿帮与抽卡总控：" + risk, "",
    ]


def _shot_production_control(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    visual = _contract_text(
        metadata, "aesthetic_priority",
        ("visual_thesis", "primary_eye_target", "secondary_visual_layer", "must_preserve", "degrade_first"), 360,
    ) or _contract_text(
        metadata, "static_aesthetic_contract",
        ("visual_intent", "composition_hierarchy", "light_design", "material_anchor", "signature_frame"), 360,
    ) or "第一视觉落点、构图焦点和一个光影/材质锚点以画面描述为准。"

    light = _contract_text(
        metadata, "lighting_topology_contract",
        ("motivated_source", "source_direction", "temperature_range", "face_light_layer", "shadow_exposure_policy", "volume_light_boundary", "conflict_resolution"), 420,
    ) or _contract_text(
        metadata, "video_texture_contract",
        ("exposure_policy", "material_motion_policy", "atmosphere_motion_policy", "continuity_carryover"), 420,
    ) or "继承场景动机光方向；脸与活动手曝光可读，高光不过曝、黑位保留层次。"

    dynamic_contract = metadata.get("dynamic_aesthetic_contract", {}) if isinstance(metadata.get("dynamic_aesthetic_contract"), dict) else {}
    dynamic_anchor = _dynamic_motion_anchor(metadata)
    dynamic = "；".join(filter(None, (
        dynamic_anchor,
        _compact(dynamic_contract.get("camera_path", "")),
        _compact(dynamic_contract.get("tempo_easing", "")),
        "失稳降级：" + _compact(dynamic_contract.get("stability_fallback", "")) if _compact(dynamic_contract.get("stability_fallback", "")) else "",
    ))) or "起幅保持可读；无新增动力源时有意静止；最终停在可继承的稳定落幅。"

    performance_contract = metadata.get("performance_contract", {}) if isinstance(metadata.get("performance_contract"), dict) else {}
    relationship = metadata.get("relationship_emotion_arc", {}) if isinstance(metadata.get("relationship_emotion_arc"), dict) else {}
    performance_parts = [
        "触发：" + _compact(performance_contract.get("trigger_event", "")) if _compact(performance_contract.get("trigger_event", "")) else "",
        "对外控制：" + _compact(performance_contract.get("display_intent", "")) if _compact(performance_contract.get("display_intent", "")) else "",
        "可见泄露：" + _compact(performance_contract.get("mask_leak", "")) if _compact(performance_contract.get("mask_leak", "")) else "",
        "身体/视线：" + "；".join(filter(None, (_compact(performance_contract.get("primary_body_action", "")), _compact(performance_contract.get("eye_focus", ""))))) if any(_compact(performance_contract.get(field, "")) for field in ("primary_body_action", "eye_focus")) else "",
        "关系转折：" + _compact(relationship.get("turn_trigger", "")) if _compact(relationship.get("turn_trigger", "")) else "",
        "余波：" + (_compact(relationship.get("shared_residue", "")) or _compact(performance_contract.get("end_residue", ""))) if (_compact(relationship.get("shared_residue", "")) or _compact(performance_contract.get("end_residue", ""))) else "",
    ]
    performance = "；".join(part for part in performance_parts if part) or "本镜无可见人物表演；情绪由当前物件、环境变化或声音余波承接。"

    continuity = _contract_text(
        metadata, "continuity_contract",
        ("start_anchor", "position_continuity", "eyeline_continuity", "prop_state", "end_anchor", "next_carryover"), 300,
    )
    physical = []
    for key, fields in (
        ("prop_lifecycle_contract", ("prop", "start_location", "contact_mode", "motion_path", "end_location", "end_orientation")),
        ("prop_functional_surface_contract", ("prop", "functional_surface", "user_view_relation", "camera_visible_surface", "grip_contact", "orientation_lock")),
        ("perspective_scale_contract", ("subjects_depth", "support_plane", "body_ratio_lock", "grounding_evidence", "motion_scaling")),
    ):
        value = _contract_text(metadata, key, fields, 260)
        if value:
            physical.append(value)
    anti_fault = _shorten("；".join(filter(None, (continuity, *physical))), 480) or "锁定人物身份、屏幕左右、身体支撑面、道具归属与稳定终态。"

    reroll = metadata.get("reroll_control", {}) if isinstance(metadata.get("reroll_control"), dict) else {}
    steps = reroll.get("mitigation_steps", []) if isinstance(reroll.get("mitigation_steps"), list) else []
    risk = "风险%s；来源：%s；策略：%s；人工首轮检查：%s" % (
        str(reroll.get("risk_level", "low") or "low"),
        _compact(reroll.get("risk_reason", "")) or "常规身份、动作与空间稳定",
        "；".join(_compact(step) for step in steps if _compact(step)) or "保持单一主体动作、单一路径和稳定终态",
        "需要" if reroll.get("manual_first_pass_check") is True else "按常规复核",
    )

    cut = metadata.get("cut_decision_contract", {}) if isinstance(metadata.get("cut_decision_contract"), dict) else {}
    temporal = metadata.get("temporal_transition_contract", {}) if isinstance(metadata.get("temporal_transition_contract"), dict) else {}
    temporal_kind = str(temporal.get("kind", "none") or "none")
    montage = "%s；切点：%s；信息增量：%s；声音承接：%s；降级：%s" % (
        "非蒙太奇" if temporal_kind in {"", "none"} else "时空/蒙太奇类型" + temporal_kind,
        _compact(cut.get("trigger", "")) or _compact(cut.get("cut_mode", "")) or "保持当前连续镜头",
        _compact(cut.get("information_gain", "")) or "保持当前唯一叙事节拍",
        _compact(cut.get("sound_strategy", "")) or _compact(temporal.get("sound_bridge", "")) or "无额外声桥",
        _compact(cut.get("fallback", "")) or _compact(temporal.get("fallback", "")) or "不稳定时保持当前构图",
    )

    return "\n".join([
        "画面质感：" + _shorten(visual, 480),
        "光效与曝光：" + _shorten(light, 480),
        "动态美学：" + _shorten(dynamic, 480),
        "表演与情绪：" + _shorten(performance, 480),
        "穿帮控制：" + anti_fault,
        "抽卡策略：" + _shorten(risk, 480),
        "蒙太奇与剪辑：" + _shorten(montage, 480),
    ])


def _scene_state_lines(master_package, plan):
    rows = []
    seen = set()
    planned_scene = {item.get("shot_id", ""): item.get("scene", "场景") for item in plan.get("shots", []) if isinstance(item, dict)}
    for shot in master_package.get("shots", []):
        if not isinstance(shot, dict):
            continue
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
        continuity = metadata.get("continuity_contract", {}) if isinstance(metadata.get("continuity_contract"), dict) else {}
        scene = planned_scene.get(shot.get("shot_id", ""), "场景")
        space_id = _compact(palette.get("space_id", "")) or scene
        key = scene + "|" + space_id
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            "%s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s" % (
                scene,
                space_id,
                _compact(palette.get("space_master_sentence", "")) or "空间主锁定见画面描述",
                _compact(palette.get("tone_palette", "")) or _compact(palette.get("visual_scene_prefix", "")) or "影调按全局风格",
                _compact(palette.get("light_texture_purpose", "")) or "光源/材质按本镜任务",
                "；".join(filter(None, (
                    "前景：" + _compact(palette.get("foreground_layer", "")) if _compact(palette.get("foreground_layer", "")) else "",
                    "中景：" + _compact(palette.get("midground_layer", "")) if _compact(palette.get("midground_layer", "")) else "",
                    "后景：" + _compact(palette.get("background_layer", "")) if _compact(palette.get("background_layer", "")) else "",
                ))) or "场景层次见画面描述",
                "；".join(filter(None, (
                    _compact(palette.get("genre_visual_signature", "")),
                    _compact(palette.get("lived_in_detail", "")),
                ))) or "题材与生活质感见画面描述",
                _compact(palette.get("depth_focus_policy", "")) or "主体实焦、背景退后",
                "；".join(filter(None, (
                    _compact(palette.get("landscape_identity", "")),
                    _compact(palette.get("landscape_composition", "")),
                ))) or "风景身份与构图见画面描述",
                "；".join(filter(None, (
                    _compact(palette.get("natural_motion_system", "")),
                    _compact(palette.get("environment_story_arc", "")),
                    _compact(palette.get("reveal_order", "")),
                    _compact(palette.get("light_weather_progression", "")),
                    _compact(palette.get("breathing_policy", "")),
                ))) or "环境演进与镜头呼吸见画面描述",
                _compact(continuity.get("prop_state", "")) or "无活动道具状态变化",
            )
        )
    if not rows:
        return ["场景 | 空间ID | 空间主锁定句 | 影调色卡句 | 固定锚点和光线 | 前中后景层次 | 题材与生活质感 | 虚实主次 | 风景身份与构图 | 环境演进与呼吸 | 活动道具/状态", ""]
    return ["场景 | 空间ID | 空间主锁定句 | 影调色卡句 | 固定锚点和光线 | 前中后景层次 | 题材与生活质感 | 虚实主次 | 风景身份与构图 | 环境演进与呼吸 | 活动道具/状态", *rows, ""]


def _global_negative_prompt(master_package):
    terms = [
        "五官漂移", "换脸", "脸型变形", "发型错乱", "服装变色", "手指畸形",
        "肢体穿模", "多手多臂", "非说话者口型乱动", "口型错位", "嘴部崩坏",
        "背景重构", "人物瞬移", "站位互换", "道具漂浮穿手", "画面跳帧",
        "过度磨皮", "模糊失焦", "人物僵硬", "全身静止", "无眨眼",
        "空洞呆滞眼神", "面部无任何变化", "肢体不动", "木偶式静止",
        "人物忽高忽低", "体型动态变化", "腿部拉长缩短", "无因尺度跳变",
        "无因浮空", "透视错乱", "穿模", "肢体畸形", "广角畸变",
    ]
    for shot in master_package.get("shots", []):
        if not isinstance(shot, dict):
            continue
        for term in re.split(r"[，,；;\n]+", str(shot.get("negative_prompt", "") or "")):
            term = term.strip()
            if term and term not in terms:
                terms.append(term)
    return "，".join(terms[:48])


def _visible_people(metadata):
    roles = metadata.get("performance_priority", {}) if isinstance(metadata, dict) else {}
    if not isinstance(roles, dict):
        return "无明确人物。"
    people = []
    if roles.get("primary"):
        people.append(str(roles.get("primary")))
    people.extend(str(item) for item in roles.get("supporting", []) if str(item).strip())
    people.extend(str(item) for item in roles.get("background", []) if str(item).strip())
    return "\n".join(dict.fromkeys(people)) if people else "无明确人物。"


def _picture_parameter_line(task, plan):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
    cinematic = metadata.get("cinematic_image_contract", {}) if isinstance(metadata.get("cinematic_image_contract"), dict) else {}
    parts = [
        "画幅：%s" % plan.get("canvas", ""),
        "风格：%s" % plan.get("visual_style", ""),
        "影调：%s" % (_compact(palette.get("tone_palette", "")) or "按场景色卡"),
        "色卡：%s" % (_compact(palette.get("visual_scene_prefix", "")) or "按视觉场景前缀"),
    ]
    if cinematic:
        parts.append("质感合同：%s" % _shorten("; ".join(str(cinematic.get(field, "")) for field in ("exposure_contrast", "color_separation", "material_detail") if cinematic.get(field)), 160))
    return "；".join(part for part in parts if part)


def _camera_description(camera_beats, task):
    if camera_beats:
        parts = []
        for beat in camera_beats[:3]:
            if not isinstance(beat, dict):
                continue
            parts.append("%s：%s，%s，%s" % (
                beat.get("time_range", ""),
                beat.get("camera_response", ""),
                beat.get("framing", ""),
                beat.get("camera_movement", ""),
            ))
        if parts:
            return "；".join(parts)
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    mode = metadata.get("editorial_mode", "continuous_take")
    return "continuous_take" if mode == "continuous_take" else str(mode)


def _lighting_description(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
    video_texture = metadata.get("video_texture_contract", {}) if isinstance(metadata.get("video_texture_contract"), dict) else {}
    cinematic = metadata.get("cinematic_image_contract", {}) if isinstance(metadata.get("cinematic_image_contract"), dict) else {}
    pieces = [
        palette.get("light_texture_purpose", ""),
        cinematic.get("exposure_contrast", ""),
        cinematic.get("atmosphere_layer", ""),
        video_texture.get("exposure_policy", ""),
        video_texture.get("material_motion_policy", ""),
    ]
    text = "；".join(_compact(piece) for piece in pieces if _compact(piece))
    return text or _shorten(_prompt_sections(task.get("full_prompt", "")).get("光照、声音与稳定约束", ""), 240)


def _state_carryover(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    continuity = metadata.get("continuity_contract", {}) if isinstance(metadata.get("continuity_contract"), dict) else {}
    return _compact(
        continuity.get("next_carryover")
        or continuity.get("end_anchor")
        or metadata.get("end_state", "")
        or "本镜结束状态见画面描述落幅。"
    )


def _cinematic_direct_clause(metadata):
    contract = metadata.get("cinematic_image_contract", {}) if isinstance(metadata, dict) else {}
    contract = contract if isinstance(contract, dict) else {}
    aesthetic = metadata.get("static_aesthetic_contract", {}) if isinstance(metadata, dict) else {}
    aesthetic = aesthetic if isinstance(aesthetic, dict) else {}
    parts = []
    anchor = _static_visual_anchor(metadata)
    if anchor:
        parts.append(anchor)
    for field in (
        "visual_intent", "composition_hierarchy", "light_design", "color_grade",
        "lens_rendering", "depth_atmosphere", "material_anchor", "signature_frame",
    ):
        value = _compact(aesthetic.get(field, ""))
        if value and value not in anchor:
            parts.append(value)
    for field in (
        "composition_anchor",
        "lens_depth",
        "exposure_contrast",
        "color_separation",
        "atmosphere_layer",
        "material_detail",
        "imperfection_map",
        "signature_frame",
    ):
        value = _compact(contract.get(field, ""))
        if value:
            parts.append(value)
    if not parts:
        return ""
    return "写实影像约束（静态美术）：" + "；".join(
        part if part == anchor else _shorten(part, 72) for part in parts[:8]
    )


def _video_texture_direct_clause(metadata, static_anchor=""):
    contract = metadata.get("video_texture_contract", {}) if isinstance(metadata, dict) else {}
    contract = contract if isinstance(contract, dict) else {}
    aesthetic = metadata.get("dynamic_aesthetic_contract", {}) if isinstance(metadata, dict) else {}
    aesthetic = aesthetic if isinstance(aesthetic, dict) else {}
    parts = []
    anchor = _dynamic_motion_anchor(metadata)
    if anchor:
        parts.append(anchor)
    if static_anchor:
        parts.append("静态锚点" + static_anchor)
    for field in (
        "motion_thesis", "primary_subject_motion", "secondary_environment_motion",
        "camera_path", "focus_behavior", "material_motion", "atmosphere_motion",
        "tempo_easing", "end_state",
    ):
        value = _compact(aesthetic.get(field, ""))
        if value and value not in anchor and value not in static_anchor:
            parts.append(value)
    for field in (
        "exposure_policy",
        "material_motion_policy",
        "atmosphere_motion_policy",
        "camera_stability_policy",
        "continuity_carryover",
    ):
        value = _compact(contract.get(field, ""))
        if value:
            parts.append(value)
    if not parts:
        return ""
    preserved = {value for value in (anchor, static_anchor) if value}
    return "视频质感约束（动态美术）：" + "；".join(
        part if part in preserved else _shorten(part, 72) for part in parts[:8]
    )


def _dynamic_motion_anchor(metadata):
    """Compile the existing dynamic contract into one causal, required sentence."""
    contract = metadata.get("dynamic_aesthetic_contract", {}) if isinstance(metadata, dict) else {}
    contract = contract if isinstance(contract, dict) else {}
    start = _compact(contract.get("start_state", ""))
    trigger = _compact(contract.get("trigger", ""))
    primary = _compact(contract.get("primary_subject_motion", ""))
    end = _compact(contract.get("end_state", ""))
    response = ""
    for field in ("secondary_environment_motion", "material_motion", "atmosphere_motion"):
        candidate = _compact(contract.get(field, ""))
        if candidate and not re.search(r"^(?:无|没有|不增加|保持静止|保持稳定|固定|none|n/a)", candidate, re.I):
            response = candidate
            break
    if not any((trigger, primary, end)):
        return ""
    parts = []
    if start:
        parts.append(start if re.search(r"^(?:起幅|起态|固定起点)", start) else "起幅" + start)
    if trigger:
        parts.append(trigger if re.search(r"^(?:因|由|被|当|听见|看见)", trigger) else "因" + trigger)
    if primary:
        parts.append(primary if re.search(r"^(?:主体|人物|角色|镜头|摄影机)", primary) else "主体" + primary)
    if response:
        parts.append(response if re.search(r"^(?:随后|同时|稍后|继而|接着)", response) else "随后" + response)
    if end:
        parts.append(end if re.search(r"^(?:最终|终态|最后|落幅|停在|停稳)", end) else "最终" + end)
    return "，".join(parts) + "。"


def _static_visual_anchor(metadata):
    """Keep one concrete light/color/material anchor available to the direct prompt."""
    contract = metadata.get("static_aesthetic_contract", {}) if isinstance(metadata, dict) else {}
    contract = contract if isinstance(contract, dict) else {}
    if not any(str(contract.get(field, "") or "").strip() for field in ("light_design", "color_grade", "material_anchor")):
        bible = metadata.get("visual_bible", {}) if isinstance(metadata, dict) else {}
        bible = bible if isinstance(bible, dict) else {}
        contract = {
            "light_design": bible.get("light_motivation", ""),
            "color_grade": bible.get("palette_system", ""),
            "material_anchor": bible.get("material_world", ""),
        }
    parts = []
    for field in ("light_design", "color_grade", "material_anchor"):
        value = _compact(contract.get(field, ""))
        if value:
            parts.append(_shorten(value, 58))
    return "，".join(parts)


def _dialogue_required_fragments(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    control = task.get("generation_control", {}) if isinstance(task.get("generation_control"), dict) else {}
    return [
        str(event.get("text", "") or "").strip()
        for event in metadata.get("dialogue_events", []) if isinstance(event, dict)
        if control.get("audio_enabled") is True and str(event.get("text", "") or "").strip()
    ]


def _budget_render_units(budget):
    value = str(budget.get("must_render", "") or "") if isinstance(budget, dict) else ""
    return [unit.strip() for unit in re.split(r"[;；|\n]+", value) if len(unit.strip()) >= 3]


def _unique_nonempty(values):
    seen, result = set(), []
    for value in values:
        value = str(value or "").strip()
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _clean_export_direct_text(text):
    text = jimeng_feed_prompt(text)
    text = re.sub(r"\s+", " ", text).strip()
    return re.sub(r"。{2,}", "。", text)


def _high_risk_direct_blocks_enabled(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    reroll = metadata.get("reroll_control", {}) if isinstance(metadata.get("reroll_control"), dict) else {}
    continuity = metadata.get("continuity_contract", {}) if isinstance(metadata.get("continuity_contract"), dict) else {}
    events = metadata.get("dialogue_events", []) if isinstance(metadata.get("dialogue_events"), list) else []
    long_dialogue = any(len(str(event.get("text", "") or "")) >= 32 for event in events if isinstance(event, dict))
    return bool(
        reroll.get("risk_level") == "high"
        or reroll.get("manual_first_pass_check") is True
        or metadata.get("editorial_mode") == "shot_group"
        or continuity.get("state_change") is True
        or long_dialogue
    )


def _build_direct_constraint_block(task):
    metadata = task.get("qa_metadata", {}) if isinstance(task.get("qa_metadata"), dict) else {}
    palette = metadata.get("scene_tone_palette", {}) if isinstance(metadata.get("scene_tone_palette"), dict) else {}
    continuity = metadata.get("continuity_contract", {}) if isinstance(metadata.get("continuity_contract"), dict) else {}
    screen_policy = metadata.get("screen_text_policy", {}) if isinstance(metadata.get("screen_text_policy"), dict) else {}
    reroll = metadata.get("reroll_control", {}) if isinstance(metadata.get("reroll_control"), dict) else {}
    pieces = [
        palette.get("space_id", ""),
        palette.get("space_master_sentence", ""),
        continuity.get("start_anchor", ""),
        continuity.get("prop_state", ""),
        continuity.get("next_carryover") or continuity.get("end_anchor", ""),
    ]
    if screen_policy.get("mode") in {"ai_overlay", "ai_generated", "ai_ui"}:
        pieces.extend([
            screen_policy.get("render_rule", ""),
            screen_policy.get("safe_area", ""),
            screen_policy.get("perspective_rule", ""),
        ])
    steps = reroll.get("mitigation_steps", []) if isinstance(reroll.get("mitigation_steps"), list) else []
    pieces.extend(str(step) for step in steps[:2])
    text = "；".join(_compact(piece) for piece in pieces if _compact(piece))
    return _shorten(_clean_export_direct_text(text), 260)


def _build_direct_negative_block(task):
    negative = str(task.get("negative_prompt", "") or "")
    terms = []
    for term in re.split(r"[，,；;\n]+", negative):
        term = term.strip()
        if term and term not in terms:
            terms.append(term)
    return "，".join(terms[:8])


def _build_master_prompt(children, plan):
    """Render one Jimeng task from one-to-three validated internal beats.

    The internal package retains per-subshot provenance for retries, while the
    Markdown delivery is deliberately one main-shot task.  This prevents users
    from accidentally submitting individual reverse shots as unrelated videos.
    """
    children = children[:3]
    sections = [_prompt_sections(item.get("full_prompt", "")) for item in children]
    first = sections[0] if sections else {}
    spatial = "；".join(_compact(section.get("主体与空间锁定", "")) for section in sections if section.get("主体与空间锁定"))
    continuity = "；".join(_compact(section.get("主镜头连续规则", "")) for section in sections if section.get("主镜头连续规则"))
    offset = 0.0
    beats = []
    for index, (child, section) in enumerate(zip(children, sections), 1):
        duration = float(child.get("duration", 0) or 0)
        content = _offset_ranges(section.get("子镜头组", ""), offset)
        content = _strip_inner_shot_headings(content)
        beats.append("【镜头%d｜%.1f-%.1f秒｜%s】%s" % (
            index, offset, offset + duration, child.get("subshot_id", ""), _compact(content)
        ))
        offset += duration
    spec = _compact(first.get("生成规格", "")) or "即梦 T2V；%s画幅；%s。" % (plan.get("canvas", ""), plan.get("visual_style", ""))
    light = "；".join(_compact(section.get("光照、声音与稳定约束", "")) for section in sections if section.get("光照、声音与稳定约束"))
    return "\n\n".join([
        "生成规格：" + spec,
        "主体与空间锁定：" + spatial,
        "主镜头连续规则：同一戏剧目标、同一场景光源和人物关系；" + continuity,
        "子镜头组：" + "\n".join(beats),
        "光照、声音与稳定约束：" + light,
    ])


def _prompt_sections(text):
    labels = ("生成规格", "主体与空间锁定", "主镜头连续规则", "子镜头组", "光照、声音与稳定约束")
    matches = list(re.finditer(r"(?:^|\n\n)(%s)[：:]" % "|".join(map(re.escape, labels)), str(text or "").strip()))
    return {match.group(1): str(text)[match.end():matches[index + 1].start() if index + 1 < len(matches) else len(str(text))].strip() for index, match in enumerate(matches)}


def _offset_ranges(text, offset):
    return re.sub(r"(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)秒", lambda m: "%.1f-%.1f秒" % (float(m.group(1)) + offset, float(m.group(2)) + offset), str(text or ""))


def _strip_inner_shot_headings(text):
    return re.sub(r"【镜头\d+[^】]*】", "", str(text or "")).strip()


def _compact(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _merge_negative_prompts(children):
    terms = []
    for child in children:
        for term in re.split(r"[，,；;\n]+", str(child.get("negative_prompt", "") or "")):
            term = term.strip()
            if term and term not in terms:
                terms.append(term)
    return "，".join(terms[:8])


def _append_execution_beats(lines, director_item):
    beats = director_item.get("camera_beat_map", []) if isinstance(director_item, dict) else []
    if not isinstance(beats, list) or not beats:
        lines.extend(["连续镜头，无额外切换。", ""])
        return
    lines.extend(["| 时间窗 | 表演触发 | 视觉主体与落幅 | 镜头响应 | 状态承接 |", "|---|---|---|---|---|"])
    for beat in beats[:3]:
        if not isinstance(beat, dict):
            continue
        lines.append("| %s | %s | %s | %s | %s |" % (
            _md_cell(beat.get("time_range", "")),
            _md_cell(beat.get("trigger", "")),
            _md_cell("%s；%s" % (beat.get("focus_subject", ""), beat.get("framing", ""))),
            _md_cell(beat.get("camera_response", "")),
            _md_cell(beat.get("carryover", "")),
        ))
    lines.append("")


def _build_transition_prompt(current_shot, next_shot):
    if not next_shot:
        return "无，段落结束。"
    current_meta = current_shot.get("qa_metadata", {}) if isinstance(current_shot.get("qa_metadata"), dict) else {}
    next_meta = next_shot.get("qa_metadata", {}) if isinstance(next_shot.get("qa_metadata"), dict) else {}
    current_contract = current_meta.get("continuity_contract", {}) if isinstance(current_meta.get("continuity_contract"), dict) else {}
    next_contract = next_meta.get("continuity_contract", {}) if isinstance(next_meta.get("continuity_contract"), dict) else {}
    current_end = _shorten(
        current_contract.get("next_carryover") or current_contract.get("end_anchor") or current_meta.get("end_state", "")
    )
    next_start = _shorten(next_contract.get("start_anchor") or next_meta.get("start_state", ""))
    transition_type = _detect_transition_type(current_contract, next_contract, current_shot, next_shot)
    body = [
        f"转场类型：{transition_type}。",
        f"上一镜落幅：{current_end}" if current_end else "上一镜落幅：保持上一镜可见残留。",
        f"下一镜开头：{next_start}" if next_start else "下一镜开头：继承上一镜残留，不复位。",
    ]
    return _shorten(" ".join(body), 180)


def _detect_transition_type(current_contract, next_contract, current_shot, next_shot):
    text = " ".join(str(v or "") for v in (
        current_contract.get("prop_state", ""),
        current_contract.get("next_carryover", ""),
        next_contract.get("start_anchor", ""),
        next_contract.get("eyeline_continuity", ""),
        next_contract.get("lighting_continuity", ""),
        current_shot.get("full_prompt", ""),
        next_shot.get("full_prompt", ""),
    ))
    if any(token in text for token in ("门", "把手", "手机", "外套", "道具", "纸", "刀", "枪", "杯", "领口", "衣角")):
        return "道具接"
    if any(token in text for token in ("视线", "目光", "看向", "回望", "对视")):
        return "视线接"
    if any(token in text for token in ("光", "亮", "暗", "色温", "阴影", "背光", "逆光", "侧光")):
        return "光线接"
    if any(token in text for token in ("说", "台词", "OS", "OV", "声音", "脚步", "关门", "呼吸")):
        return "声桥"
    if any(token in text for token in ("走", "转身", "抬手", "停住", "落步", "前倾", "后仰", "推进", "拉回")):
        return "动作接"
    if any(token in text for token in ("同框", "构图", "站位", "位置", "距离", "屏幕方向")):
        return "同构图接"
    return "硬切"


def _shorten(text, limit=120):
    text = str(text or "").replace("\r", " ").replace("\n", " ").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip("，。；; ") + "…"


def _write_workbook(path, package, plan, director):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    shots = package.get("shots", [])
    director_map = {
        item.get("subshot_id", ""): item for item in director.get("items", []) if item.get("subshot_id")
    }
    workbook = Workbook()
    prompts = workbook.active
    prompts.title = "AI视频模型提示词"
    prompts.append([
        "主镜头", "子镜头", "时长(s)", "画面描述｜直接复制", "负面提示词",
        "生成模式", "原生音频", "人工首轮验证",
    ])
    for shot in shots:
        control = shot.get("generation_control", {}) if isinstance(shot.get("generation_control"), dict) else {}
        prompts.append([
            shot.get("shot_id", ""),
            shot.get("subshot_id", ""),
            shot.get("duration", 0),
            _build_direct_copy_prompt(shot, plan),
            shot.get("negative_prompt", ""),
            control.get("mode", ""),
            control.get("audio_enabled", False),
            (shot.get("qa_metadata", {}).get("reroll_control", {}) or {}).get("manual_first_pass_check", False),
        ])

    qa = workbook.create_sheet("QA与表演预算")
    qa.append([
        "主镜头", "子镜头", "戏剧目标", "角色场景目标/策略", "关系情绪弧",
        "序列导演计划", "剪辑切点", "提示词信息预算", "声音导演计划", "道具功能面合同", "肤色保护合同", "主表演者", "对手", "背景",
        "通用道具生命周期", "透视比例保护", "光源拓扑",
        "镜头功能", "叙事权重", "信息增量", "反应归属", "戏剧节拍ID",
        "时长策略", "内容所需时长", "容量利用率", "时长依据",
        "主动作数", "情绪转折数", "对手反应数", "实体运镜数", "镜头响应数", "起始状态", "终态",
        "表演因果", "表演张力合同", "戏眼合同", "连续性合同", "抽卡控制", "台词引用", "注意力交接", "打斗连续性",
    ])
    for shot in shots:
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        roles = metadata.get("performance_priority", {}) if isinstance(metadata.get("performance_priority"), dict) else {}
        budget = metadata.get("action_budget", {}) if isinstance(metadata.get("action_budget"), dict) else {}
        dramatic = metadata.get("dramatic_design", {}) if isinstance(metadata.get("dramatic_design"), dict) else {}
        duration_design = metadata.get("duration_design", {}) if isinstance(metadata.get("duration_design"), dict) else {}
        qa.append([
            shot.get("shot_id", ""), shot.get("subshot_id", ""), metadata.get("dramatic_goal", ""),
            json.dumps(metadata.get("character_scene_objective_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("relationship_emotion_arc", {}), ensure_ascii=False),
            json.dumps(metadata.get("sequence_directing_plan", {}), ensure_ascii=False),
            json.dumps(metadata.get("cut_decision_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("prompt_information_budget", {}), ensure_ascii=False),
            json.dumps(metadata.get("sound_directing_plan", {}), ensure_ascii=False),
            json.dumps(metadata.get("prop_functional_surface_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("skin_tone_protection_contract", {}), ensure_ascii=False),
            roles.get("primary", ""), "；".join(roles.get("supporting", [])), "；".join(roles.get("background", [])),
            json.dumps(metadata.get("prop_lifecycle_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("perspective_scale_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("lighting_topology_contract", {}), ensure_ascii=False),
            dramatic.get("shot_function", ""), dramatic.get("narrative_weight", ""),
            dramatic.get("information_gain", ""), dramatic.get("reaction_ownership", ""),
            "；".join(dramatic.get("dramatic_beat_ids", [])),
            duration_design.get("duration_strategy", ""), duration_design.get("justified_content_duration", ""),
            duration_design.get("utilization_ratio", ""), duration_design.get("duration_rationale", ""),
            budget.get("primary_action_count", 0), budget.get("emotion_turn_count", 0),
            budget.get("supporting_reaction_count", 0), budget.get("physical_camera_move_count", 0), budget.get("editorial_response_count", 0),
            metadata.get("start_state", ""), metadata.get("end_state", ""),
            json.dumps(metadata.get("performance_causality", {}), ensure_ascii=False),
            json.dumps(metadata.get("performance_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("story_punch_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("continuity_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("reroll_control", {}), ensure_ascii=False),
            "；".join(metadata.get("dialogue_refs", [])),
            json.dumps(metadata.get("attention_handoff", {}), ensure_ascii=False),
            json.dumps(metadata.get("fight_continuity", {}), ensure_ascii=False),
        ])

    dialogue_sheet = workbook.create_sheet("台词与OS表演")
    dialogue_sheet.append([
        "主镜头", "子镜头", "引用", "类型", "人物", "逐字原文", "时间窗",
        "人物可见性", "台词功能", "潜台词", "原文重音词", "潜台词可见证据", "轮次关系",
        "会话模式", "响应延迟", "抢话/打断窗口", "会话源文依据",
        "发声时神态", "发声时身体状态", "语气与停顿", "气口计划", "口型同步", "原生音频",
    ])
    for shot in shots:
        metadata = shot.get("qa_metadata", {}) if isinstance(shot.get("qa_metadata"), dict) else {}
        control = shot.get("generation_control", {}) if isinstance(shot.get("generation_control"), dict) else {}
        for event in metadata.get("dialogue_events", []) if isinstance(metadata.get("dialogue_events"), list) else []:
            dialogue_sheet.append([
                shot.get("shot_id", ""), shot.get("subshot_id", ""), event.get("ref", ""),
                event.get("kind", ""), event.get("speaker", ""), event.get("text", ""),
                event.get("time_range", ""), event.get("speaker_visibility", ""),
                event.get("line_function", ""), event.get("subtext", ""),
                "；".join(str(word) for word in event.get("stress_words", []) if str(word).strip()),
                event.get("subtext_visible_evidence", ""), event.get("turn_relation", ""),
                event.get("conversation_mode", ""), event.get("response_latency", ""),
                event.get("overlap_or_interrupt_window", ""), event.get("conversation_source_basis", ""),
                event.get("facial_state", ""), event.get("body_state", ""), event.get("delivery", ""), event.get("breath_pause_plan", ""),
                event.get("lip_sync", False), control.get("audio_enabled", False),
            ])

    director_sheet = workbook.create_sheet("导演连续性")
    director_sheet.append([
        "主镜头", "子镜头", "景别", "机位", "运镜", "视点", "画面层级",
        "入场策略", "揭示策略", "焦点策略", "镜头模式", "表演链", "镜头执行节拍",
        "序列承接", "轴线", "灯光", "落幅",
    ])
    for shot in shots:
        item = director_map.get(shot.get("subshot_id", ""), {})
        director_sheet.append([
            shot.get("shot_id", ""), shot.get("subshot_id", ""), item.get("shot_size", ""),
            item.get("camera_position", item.get("camera_relative_pos", "")),
            item.get("movement_detail", item.get("camera", "")), item.get("viewpoint", ""),
            item.get("visual_hierarchy", ""), item.get("entry_strategy", ""),
            item.get("reveal_strategy", ""), item.get("focus_strategy", ""),
            item.get("editorial_mode", ""),
            json.dumps(item.get("performance_chain", {}), ensure_ascii=False),
            json.dumps(item.get("camera_beat_map", []), ensure_ascii=False),
            json.dumps(item.get("sequence_context", {}), ensure_ascii=False), item.get("axis_space", ""),
            item.get("lighting", ""), item.get("end_state", ""),
        ])

    plan_by_shot = {item.get("shot_id", ""): item for item in plan.get("shots", []) if isinstance(item, dict)}
    keyframe_sheet = workbook.create_sheet("关键帧流水线")
    keyframe_sheet.append([
        "主镜头", "子镜头", "优先级", "触发原因", "帧类型", "时间(s)",
        "关键帧生图提示词", "即梦视频提示｜配合关键帧", "人物/道具状态差异",
        "关键帧连续性检查", "关键帧-T2V事实一致性", "负面提示词",
    ])
    for shot in shots:
        planned = plan_by_shot.get(shot.get("shot_id", ""), {})
        sequence = build_keyframe_sequence(shot, planned, plan.get("canvas", "16:9"), plan.get("visual_style", ""))
        if not sequence:
            continue
        state_diff = json.dumps(sequence.get("state_diff", []), ensure_ascii=False)
        continuity_check = json.dumps(sequence.get("continuity_check", []), ensure_ascii=False)
        fact_consistency = json.dumps(sequence.get("fact_consistency", []), ensure_ascii=False)
        for frame in sequence["frames"]:
            keyframe_sheet.append([
                shot.get("shot_id", ""), shot.get("subshot_id", ""), sequence.get("priority", ""),
                sequence.get("reason", ""), frame.get("label", ""), frame.get("time_seconds", 0),
                frame.get("prompt", ""), sequence.get("video_prompt", ""), state_diff,
                continuity_check, fact_consistency, sequence.get("negative_prompt", ""),
            ])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in sheet.columns:
            letter = column[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 80)
    workbook.save(path)
    return True


def _md_cell(value):
    text = str(value).replace("\r", " ").replace("\n", " ")
    return text.replace("|", "\\|")


def _dialogue_md_cell(event, field):
    value = event.get(field, "") if isinstance(event, dict) else ""
    if field == "stress_words" and isinstance(value, list):
        value = "；".join(str(word) for word in value if str(word).strip())
    return _md_cell(value)


def _find_package(run_dir):
    for relative in (
        ".cache/composer/merged.prompt_package.json",
        ".cache/composer/prompt_package.json",
        ".cache/prompt_package.json",
    ):
        path = os.path.join(run_dir, relative)
        if os.path.exists(path):
            return path
    return ""


def _load(path):
    with open(path, "r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def _load_optional(path):
    return _load(path) if os.path.exists(path) else {"items": []}


if __name__ == "__main__":
    args = [argument for argument in sys.argv[1:] if argument != "--regenerate"]
    if len(args) != 2:
        print("Usage: python3 export_with_validation.py <user_confirmed_export_md> <run_dir>")
        print("ERROR: output path is mandatory. Ask the user for the export file location before running.")
        sys.exit(2)
    sys.exit(export_with_validation(args[0], args[1]))
