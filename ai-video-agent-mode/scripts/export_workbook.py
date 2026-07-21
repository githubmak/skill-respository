"""Mode C v4 Excel export with model/negative/QA/control separation."""
import json, os, re, sys
if not os.environ.get("PYTHONPYCACHEPREFIX") and not getattr(sys, "pycache_prefix", None):
    sys.dont_write_bytecode = True
sys.path.insert(0, os.path.dirname(__file__))
from pycache_policy import block_source_pycache_until_run_dir, ensure_pycache_prefix_from_path

block_source_pycache_until_run_dir()
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
ALT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")


def style_header(ws, row=1):
    """Apply header styling to first row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_body(ws, start_row=2):
    """Apply body styling to data rows."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            # Alternating rows
            if (cell.row - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def auto_width(ws, max_width=60, min_width=8):
    """Set reasonable column widths."""
    for i, col in enumerate(ws.columns, 1):
        lengths = []
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                lengths.append(max(len(l) for l in lines))
        best = max(lengths) if lengths else min_width
        ws.column_dimensions[get_column_letter(i)].width = min(max(best + 2, min_width), max_width)


def _field(item, *keys, default=""):
    for key in keys:
        val = item.get(key)
        if val not in (None, ""):
            return val
    return default


def _prompt_section(full_prompt, label):
    if not full_prompt:
        return ""
    pattern = rf"{re.escape(label)}[：:](.*?)(?=\n\n\S+?[：:]|$)"
    m = re.search(pattern, full_prompt, flags=re.S)
    return m.group(1).strip() if m else ""


def _derived_field(item, field_name):
    fp = str(item.get("full_prompt", ""))
    if field_name == "shot_size":
        m = re.search(r"景别[：:]\s*([^。；;\n]+)", fp)
        return m.group(1).strip() if m else ""
    if field_name == "camera_position":
        m = re.search(r"机位[：:]\s*([^。；;\n]+)", fp)
        return m.group(1).strip() if m else ""
    if field_name == "camera":
        m = re.search(r"运镜描述[：:]\s*([^。；;\n]+)", fp)
        return m.group(1).strip() if m else ""
    if field_name == "lighting":
        return _prompt_section(fp, "光照与声音")
    if field_name == "end_state":
        metadata = item.get("qa_metadata", {})
        return metadata.get("end_state", "") if isinstance(metadata, dict) else ""
    if field_name == "character_action":
        return _prompt_section(fp, "表演时间轴")
    return ""


def export(pkg_path, plan_path, out):
    """Generate styled 7-sheet Excel workbook."""
    ensure_pycache_prefix_from_path(out)
    out_dir = os.path.dirname(os.path.abspath(out))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(pkg_path, "r", encoding="utf-8-sig") as f:
        pp = json.load(f)
    with open(plan_path, "r", encoding="utf-8-sig") as f:
        sp = json.load(f)

    items = pp.get("items", [])
    merged = pp.get("merged_full_prompts", [])
    project_name = sp.get("project_name", "Untitled")

    # Group items by shot
    shot_map = {}
    for item in items:
        sid = item["shot_id"]
        shot_map.setdefault(sid, []).append(item)

    wb = Workbook()

    # ===== Sheet 1: 分镜总表 =====
    ws1 = wb.active
    ws1.title = "分镜总表"
    headers1 = ["镜头编号", "子镜头", "时长(s)", "景别", "机位", "运镜", "可见人物", "灯光", "落幅"]
    ws1.append(headers1)
    for item in items:
        ws1.append([
            _field(item, "shot_id"), _field(item, "subshot_id"), _field(item, "duration", "duration_sec", default=0),
            _field(item, "shot_size", default=_derived_field(item, "shot_size")),
            _field(item, "camera_position", "camera_relative_pos", default=_derived_field(item, "camera_position")),
            _field(item, "camera", "movement_detail", "movement_description", default=_derived_field(item, "camera")),
            _field(item, "visible_characters", "characters"),
            _field(item, "lighting", default=_derived_field(item, "lighting")),
            _field(item, "end_state", default=_derived_field(item, "end_state")),
        ])
    style_header(ws1); style_body(ws1); auto_width(ws1)
    ws1.freeze_panes = "A2"

    # ===== Sheet 2: 模型提示词 =====
    ws2 = wb.create_sheet("模型提示词")
    ws2.append(["镜头编号", "子镜头", "Mode C v4模型提示词"])
    for item in items:
        ws2.append([_field(item, "shot_id"), _field(item, "subshot_id"), _field(item, "full_prompt")])
    style_header(ws2); style_body(ws2)
    ws2.column_dimensions["C"].width = 100
    ws2.freeze_panes = "A2"

    # ===== Sheet 3: 生成控制 =====
    ws3 = wb.create_sheet("生成控制")
    ws3.append(["镜头编号", "子镜头", "生成模式与参考资产"])
    for item in items:
        ws3.append([_field(item, "shot_id"), _field(item, "subshot_id"), json.dumps(item.get("generation_control", {}), ensure_ascii=False)])
    style_header(ws3); style_body(ws3)
    ws3.column_dimensions["C"].width = 100
    ws3.freeze_panes = "A2"

    # ===== Sheet 4: QA与动作预算 =====
    ws4 = wb.create_sheet("QA与动作预算")
    ws4.append(["镜头编号", "子镜头", "时长(s)", "QA元数据", "表演张力合同", "连续性合同", "抽卡控制"])
    for item in items:
        metadata = item.get("qa_metadata", {}) if isinstance(item.get("qa_metadata"), dict) else {}
        ws4.append([
            _field(item, "shot_id"),
            _field(item, "subshot_id"),
            _field(item, "duration", "duration_sec", default=0),
            json.dumps(metadata, ensure_ascii=False),
            json.dumps(metadata.get("performance_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("continuity_contract", {}), ensure_ascii=False),
            json.dumps(metadata.get("reroll_control", {}), ensure_ascii=False),
        ])
    style_header(ws4); style_body(ws4)
    ws4.column_dimensions["D"].width = 100
    ws4.column_dimensions["E"].width = 80
    ws4.column_dimensions["F"].width = 80
    ws4.column_dimensions["G"].width = 80
    ws4.freeze_panes = "A2"

    # ===== Sheet 5: 九宫格剧情分镜图 =====
    ws5 = wb.create_sheet("九宫格剧情分镜图")
    ws5.append(["分镜编号", "镜头1", "镜头2", "镜头3"])
    ws5.append(["", "", "", ""])
    ws5.append(["镜头4", "镜头5", "镜头6"])
    ws5.append(["", "", "", ""])
    ws5.append(["镜头7", "镜头8", "镜头9"])
    ws5.append(["", "", "", ""])
    # Group shots by scene into 9-panel blocks
    row = 8
    scenes = {}
    for shot in sp.get("shots", []):
        scenes.setdefault(shot.get("scene", "?"), []).append(shot["shot_id"])
    for scene, sids in scenes.items():
        ws5.append(["场景：%s" % scene, "", "", ""])
        ws5.append(["", "", "", ""])
        ws5.cell(row=row, column=1).font = Font(bold=True, size=11, color="1F4E79")
        row += 1
        for i in range(0, len(sids), 9):
            chunk = sids[i:i+9]
            grid = chunk + [""] * (9 - len(chunk))
            for r in range(3):
                ws5.append([grid[r*3] if r*3 < len(grid) else "",
                           grid[r*3+1] if r*3+1 < len(grid) else "",
                           grid[r*3+2] if r*3+2 < len(grid) else ""])
                ws5.append(["", "", ""])
                row += 2
            ws5.append(["───", "───", "───"])
            row += 1
    style_header(ws5); style_body(ws5, start_row=2)
    for c in "ABC":
        ws5.column_dimensions[c].width = 30

    # ===== Sheet 6: 四宫格关键帧展开 =====
    ws6 = wb.create_sheet("四宫格关键帧展开")
    ws6.append(["主镜头", "关键帧1", "关键帧2", "关键帧3", "关键帧4"])
    for sid, subshots in sorted(shot_map.items()):
        row_data = [sid]
        for i in range(4):
            if i < len(subshots):
                ss = subshots[i]
                row_data.append("【%s】%s\n动作：%s\n灯光：%s" % (
                    _field(ss, "subshot_id"),
                    _field(ss, "shot_size", default=_derived_field(ss, "shot_size")),
                    _field(ss, "character_action", default=_derived_field(ss, "character_action"))[:50],
                    _field(ss, "lighting", default=_derived_field(ss, "lighting"))[:50]))
            else:
                row_data.append("")
        ws6.append(row_data)
    style_header(ws6); style_body(ws6)
    for c in "BCDE":
        ws6.column_dimensions[c].width = 50
    ws6.column_dimensions["A"].width = 12
    ws6.freeze_panes = "A2"

    # ===== Sheet 7: 负面提示词 =====
    ws7 = wb.create_sheet("负面提示词")
    ws7.append(["镜头编号", "子镜头", "负面提示词"])
    for item in items:
        ws7.append([_field(item, "shot_id"), _field(item, "subshot_id"), _field(item, "negative_prompt")])
    style_header(ws7); style_body(ws7)
    ws7.column_dimensions["C"].width = 80
    ws7.freeze_panes = "A2"

    wb.save(out)
    print("[EXPORT] %s saved (%d sheets)" % (out, len(wb.sheetnames)))


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("usage: export_workbook.py <pkg.json> <plan.json> <output.xlsx>")
        sys.exit(1)
    export(sys.argv[1], sys.argv[2], sys.argv[3])
