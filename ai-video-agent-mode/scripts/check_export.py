"""Export self-review + quality gate script. Run after every export.

Usage: python check_export.py <export_md_path> <run_dir>
       python check_export.py --quality <run_dir>  # deep director_pass check

Checks 01-15: Export format validation
Checks 16-20: Quality gate (character detail, camera, axis, lighting)
"""

import json, re, os, sys

QUALITY_MODE = "--quality" in sys.argv

def check_export(md_path, run_dir, quality_mode=False):
    """Run up to 20-point quality review. Returns (passed, issues)."""
    
    if not quality_mode:
        with open(md_path, "r", encoding="utf-8") as f:
            md_text = f.read()
    else:
        md_text = ""
    
    with open(os.path.join(run_dir, ".cache", "orchestrator", "shot_plan.json"), "r", encoding="utf-8") as f:
        sp = json.load(f)
    dm = sp.get("dialogue_map", {})
    
    try:
        with open(os.path.join(run_dir, ".cache", "director", "director_pass.json"), "r", encoding="utf-8") as f:
            dp = json.load(f)
        dp_items = dp.get("items", [])
    except:
        dp_items = []

    try:
        pkg_path = os.path.join(run_dir, ".cache", "composer", "prompt_package.json")
        if not os.path.exists(pkg_path):
            # Try merged variant — Phase 6c normalization produces this
            pkg_path = os.path.join(run_dir, ".cache", "composer", "merged.prompt_package.json")
        if not os.path.exists(pkg_path):
            # Try any .prompt_package.json
            comp_dir = os.path.join(run_dir, ".cache", "composer")
            for fn in sorted(os.listdir(comp_dir)) if os.path.isdir(comp_dir) else []:
                if fn.endswith(".prompt_package.json"):
                    pkg_path = os.path.join(comp_dir, fn)
                    break
        with open(pkg_path, "r", encoding="utf-8") as f:
            pkg = json.load(f)
    except:
        pkg = {"shots": []}
    
    plan_subshots = sum(len(s.get("subshots",[])) for s in sp["shots"])
    plan_main = len(sp["shots"])
    issues = []
    
    def chk(n, label, ok, detail):
        print(f"  [{'OK' if ok else '!!'}] {n:02d}. {label}: {detail}")
        if not ok: issues.append(f"{label}: {detail}")
        return ok
    
    if not quality_mode:
        # === FORMAT CHECKS 01-15 ===
        md_main = len(re.findall(r"### S1-\d+（", md_text))
        chk(1, "Main shots", md_main >= plan_main * 0.95, f"{md_main}/{plan_main}")
        
        dur_count = len(re.findall(r"子镜头.+?\*\*时长\*\*：\d+秒", md_text))
        chk(2, "Duration labels", dur_count >= plan_subshots, f"{dur_count}/{plan_subshots}")
        
        dia_speakers = len(re.findall(r"对白声音（[^）]{1,8}）", md_text))
        os_speakers = len(re.findall(r"OS（[^）]{1,8}）", md_text))
        dia_shots = sum(1 for s in sp["shots"] for ss in s.get("subshots",[]) if ss.get("dialogue_refs",[]))
        if dia_shots == 0:
            chk(3, "Dialogue speakers", True, "N/A")
        else:
            chk(3, "Dialogue speakers", dia_speakers+os_speakers >= dia_shots*0.9, f"{dia_speakers+os_speakers}/{dia_shots}")
        
        chk(4, "Empty scene", md_text.count("场景：。")==0, str(md_text.count("场景：。")))
        chk(5, "Name consistency", sum(md_text.count(n) for n in ["沈星州"])==0, "OK")
        chk(6, "Double punct", md_text.count("。。")==0, str(md_text.count("。。")))
        
        scenes = len(re.findall(r"^## ", md_text, re.MULTILINE))
        bq = len(re.findall(r"^> \*\*", md_text, re.MULTILINE))
        chk(7, "Blockquote format", scenes>=1 and bq>=10, f"{scenes}sc/{bq}fields")
        
        plan_ids = set()
        for shot in sp["shots"]:
            for ss in shot.get("subshots", []):
                plan_ids.add(ss["subshot_id"])
        md_ids = set(re.findall(r"子镜头 (S1-\d+-\d+)", md_text))
        chk(8, "Subshot coverage", not (plan_ids-md_ids), f"{len(md_ids)}/{len(plan_ids)}")
        
        used_refs = set()
        for s in pkg.get("shots", pkg.get("items", [])):
            used_refs.update(re.findall(r"\[(D\d+(?:-MAIN)?(?:-SYS)?)\]", s.get("full_prompt","")))
        bad = sum(1 for r in used_refs if r not in dm)
        chk(9, "Dialogue refs", bad==0, f"{len(used_refs)} refs, {bad} bad")
        
        cam_light = 0
        for item in dp_items:
            f = 0
            if len(str(item.get("camera_relative_pos", item.get("camera_position","")))) > 2: f += 1
            if item.get("camera_lens_mm", 0): f += 1
            if len(str(item.get("shot_size",""))) > 1: f += 1
            if len(str(item.get("camera_height_relative",""))) > 1: f += 1
            if f < 2: cam_light += 1
        chk(10, "Camera fields", cam_light <= 10, f"{cam_light} light")
        
        beat_issues = sum(1 for item in dp_items if len(str(item.get("character_action","")))>10 and len(set(re.findall(r"(起幅|推进|落幅)", str(item.get("character_action","")))))<2)
        chk(11, "Action 3-beat", beat_issues==0, str(beat_issues))
        
        eng = sum(1 for s in pkg.get("shots", pkg.get("items", [])) if re.search(r"(?<![a-zA-Z])(CU|ECU|MCU|MS|LS|FS|ELS)(?![a-zA-Z])", s.get("full_prompt","")))
        chk(12, "English shot sizes", eng==0, str(eng))
        
        xlsx_path = md_path.replace(".md", ".xlsx")
        if os.path.exists(xlsx_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(xlsx_path, read_only=True)
                chk(13, "XLSX rows", wb.active.max_row >= plan_subshots, f"{wb.active.max_row}/{plan_subshots+1}")
            except:
                chk(13, "XLSX", False, "read error")
        else:
            chk(13, "XLSX", False, "not found")
        
        enc_ok = False
        for enc in ["utf-8","utf-8-sig"]:
            try: open(md_path,"r",encoding=enc).read(100); enc_ok=True; break
            except: pass
        chk(14, "Encoding", enc_ok, "utf-8" if enc_ok else "BAD")
        
        short = sum(1 for s in pkg.get("shots", pkg.get("items", [])) if len(s.get("full_prompt",""))<500)
        chk(15, "Prompt length", short==0, f"{short} under 500")
    
    # === QUALITY GATE CHECKS 16-20 ===
    facial_kw = ["眼睑","瞳孔","唇线","鼻翼","呼吸","眉毛","嘴角","下颌","睫毛","眼球","眼神光","面肌"]
    facial_rich = sum(1 for item in dp_items if sum(1 for kw in facial_kw if kw in str(item.get("character_action",""))) >= 3)
    chk(16, "Facial detail", facial_rich >= len(dp_items)*0.7, f"{facial_rich}/{len(dp_items)} >=70%")
    
    lens_issues = sum(1 for item in dp_items if isinstance(item.get("camera_lens_mm",0),(int,float)) and item.get("camera_lens_mm",0)>0 and item.get("camera_lens_mm",0)<=28 and str(item.get("shot_size","")) in ("特写","大特写","近景"))
    chk(17, "Lens sanity", lens_issues==0, f"{lens_issues} UW on CU")
    
    # Refined axis check: per-character position extraction
    import re as _re
    axis_flips = 0
    prev = {}
    
    def get_char_pos(axis_text, char_name):
        for pat in [_re.escape(char_name)+r"(?:在画面)?(画左|偏左)", _re.escape(char_name)+r"(?:在画面)?(画右|偏右)"]:
            m = _re.search(pat, axis_text)
            if m: return "left" if "左" in m.group(1) else "right"
        return ""
    
    def has_axis_transition(axis_text):
        return any(kw in axis_text for kw in ["越轴","过渡","绕行","不适用","特写镜头","单人镜头","正反打","过肩","转场","建立","establishing"])
    
    def get_scene(sid):
        n = int(sid.split("-")[1])
        if n <= 15: return "VIP"
        elif n <= 75: return "canteen"
        else: return "gate"
    
    for shot in sp["shots"]:
        s_scene = get_scene(shot["shot_id"])
        for ss in shot.get("subshots", []):
            ssid = ss["subshot_id"]
            di = next((i for i in dp_items if i.get("subshot_id")==ssid), {})
            axis = str(di.get("axis_space",""))
            chars = ss.get("characters",[])
            if isinstance(chars, str): chars = [c.strip() for c in chars.split(";")]
            
            if has_axis_transition(axis):
                for ch in chars:
                    prev[ch] = {"pos": get_char_pos(axis,ch), "scene": s_scene, "sid": ssid}
                continue
            
            for ch in chars:
                cur_pos = get_char_pos(axis, ch)
                if cur_pos and ch in prev and prev[ch]["scene"]==s_scene and prev[ch]["pos"]:
                    if cur_pos != prev[ch]["pos"]:
                        axis_flips += 1
                if cur_pos:
                    prev[ch] = {"pos": cur_pos, "scene": s_scene, "sid": ssid}
    
    chk(18, "Axis continuity", axis_flips <= 5, f"{axis_flips} flips (same scene)")
    
    monotony = 0
    for shot in sp["shots"]:
        subs = shot.get("subshots",[])
        if len(subs) >= 3:
            sizes = []
            for ss in subs:
                di = next((i for i in dp_items if i.get("subshot_id")==ss["subshot_id"]), {})
                sizes.append(di.get("shot_size",""))
            if len(set(s for s in sizes if s)) <= 1:
                monotony += 1
    chk(19, "Shot variety", monotony <= 2, f"{monotony} monotonous groups")
    
    light_jumps = 0
    prev_temp = 0
    for shot in sp["shots"]:
        for ss in shot.get("subshots", []):
            di = next((i for i in dp_items if i.get("subshot_id")==ss["subshot_id"]), {})
            temps = re.findall(r"(\d+)K", str(di.get("lighting","")))
            ct = int(temps[0]) if temps else 0
            if prev_temp and ct and abs(ct-prev_temp) > 2000:
                light_jumps += 1
            if ct: prev_temp = ct
    chk(20, "Lighting continuity", light_jumps <= 8, f"{light_jumps} jumps >2000K")
    
    # === CHECK 21: Push/pull speed sanity ===
    push_issues = 0
    for item in dp_items:
        movement = str(item.get("movement_detail", item.get("movement_description", "")))
        mtype = str(item.get("movement_type", ""))
        if "push" in mtype.lower() or "推" in movement or "push" in movement:
            speeds = re.findall(r"(\d+\.?\d*)\s*m/s", movement)
            for _s in speeds:
                if float(_s) > 0.3:
                    push_issues += 1
        if "pull" in mtype.lower() or "拉" in movement or "pull" in movement:
            speeds = re.findall(r"(\d+\.?\d*)\s*m/s", movement)
            for _s in speeds:
                if float(_s) > 0.2:
                    push_issues += 1
    chk(21, "Push/pull speed", push_issues <= 3, f"{push_issues} over limit")

    # === CHECK 22: Negative prompt presence ===
    neg_missing = 0
    neg_keywords = ["负面提示", "消极提示", "negative prompt", "画面崩坏", "面部扭曲", "多余肢体"]
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        has_neg = any(kw in fp for kw in neg_keywords)
        if not has_neg:
            neg_missing += 1
    chk(22, "Negative prompt", neg_missing == 0, f"{neg_missing} missing")

    # === MODE C CHECKS 23-30 ===
    # C1: No abstract emotion labels in prompt text
    abstract_emo = ["紧张地", "开心地", "难过地", "愤怒地", "害怕地", "慌张地", "悲伤地",
                    "无奈地", "坚定地", "温柔地", "冷漠地", "轻蔑地", "惊讶地", "尴尬地",
                    "得意地", "焦虑地", "烦躁地", "绝望地", "色地", "嫉妒地"]
    abstract_hits = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        # Only check content sections (before 负面提示词)
        neg_idx = fp.find("负面提示词")
        content = fp[:neg_idx] if neg_idx > 0 else fp
        for emo in abstract_emo:
            if emo in content:
                abstract_hits += 1
                break
    chk(23, "C1: No abstract emotion labels", abstract_hits == 0,
        f"{abstract_hits} shots with bare labels")

    # C2: Every visible character >=1 independent micro-reaction
    ACTION_VERBS = [
        "走向", "转身", "抬手", "迈步", "注视", "看向", "移动", "微动", "呼吸",
        "重心", "肩", "头", "手", "脚", "身体", "步伐", "抬头", "低头", "回头",
        "侧身", "弯腰", "踏步", "往前", "向前", "向后", "向左", "向右", "靠近",
        "走开", "离开", "抬腿", "站起", "坐下", "后退", "扭头", "回身", "进入",
        "走出", "挥手", "提起", "放下", "紧握", "摇头", "点头", "抬眸", "抬眉",
        "微侧", "微偏", "睁眼", "闭眼", "眨眼", "抬起", "下沉", "绷紧", "放松",
        "收缩", "扩张", "上扬", "下压", "抿紧", "微启", "咬紧", "松开", "鼓胀",
        "滚动", "眯起", "瞪大", "扫视", "打量", "盯着", "瞥", "瞧着"
    ]
    CHAR_NAMES = ["沈星洲", "沈星雨", "向云初", "江训", "陆序", "许承"]
    per_char_missing = 0
    per_char_details = []
    for shot in sp["shots"]:
        for ss in shot.get("subshots", []):
            ssid = ss["subshot_id"]
            chars = ss.get("characters", [])
            if isinstance(chars, str):
                chars = [c.strip() for c in chars.split(";")]
            # Find corresponding prompt
            ps = next((s for s in pkg.get("shots", pkg.get("items", [])) if s.get("subshot_id") == ssid), None)
            if not ps or len(chars) < 2:
                continue
            fp = ps.get("full_prompt", "")
            neg_idx = fp.find("负面提示词")
            content = fp[:neg_idx] if neg_idx > 0 else fp
            # For each character, check if they have at least one action verb
            missing = []
            for ch in chars:
                if ch not in CHAR_NAMES:
                    continue
                sentences = re.split(r"[。；\n]", content)
                char_sents = [s for s in sentences if ch in s]
                has_action = any(
                    any(verb in s for verb in ACTION_VERBS)
                    for s in char_sents
                )
                if not has_action:
                    missing.append(ch)
            if missing and len(chars) >= 2:
                per_char_missing += 1
                per_char_details.append(f"{ssid}: {missing}")
    chk(24, "C2: Every char >=1 action (multi-char shots)",
        per_char_missing <= 5,
        f"{per_char_missing} shots with frozen chars" +
        (f" [{'; '.join(per_char_details[:5])}...]" if per_char_details else ""))

    # C3: Micro-reactions contain cause chain (trigger -> face -> body -> voice)
    cause_chain_missing = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        neg_idx = fp.find("负面提示词")
        time_seg = fp[fp.find("时间分段叙事"):neg_idx] if neg_idx > 0 else fp
        # Check for cause chain pattern: character triggers and character reacts
        has_cause = bool(re.search(r"(?:因为|导致|触发|引起|使得|令|让|看到|听到|察觉|发现|感到|注意到)", time_seg))
        has_reaction = bool(re.search(r"(?:" + "|".join(ACTION_VERBS[:10]) + r")", time_seg))
        if not (has_cause and has_reaction):
            cause_chain_missing += 1
    chk(25, "C3: Cause chain in multi-char shots",
        cause_chain_missing <= 10,
        f"{cause_chain_missing} shots lacking cause chain")

    # C4: Expressions as verb/number sequences, not noun labels
    noun_label_pattern = r"(?:紧张|开心|难过|悲伤|愤怒|害怕|惊慌|得意|尴尬|焦虑|温柔|冷漠|坚定)"
    noun_label_hits = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        neg_idx = fp.find("负面提示词")
        # Check time segment for standalone emotion nouns
        time_seg = fp[fp.find("时间分段叙事"):neg_idx] if neg_idx > 0 else fp
        # Look for emotion labels that are NOT part of a verb chain
        standalone = re.findall(r"[。；](?:[^。；]{0,5})(" + noun_label_pattern + r")(?:[^。；]{0,5})[。；]", time_seg)
        if standalone:
            noun_label_hits += 1
    chk(26, "C4: Expressions as verb/num sequences",
        noun_label_hits <= 8,
        f"{noun_label_hits} shots with noun labels")

    # C5: Scene purpose sentence present
    purpose_missing = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        # Look for "此镜" or "戏剧目标" in block 3 (时长运镜场景目的)
        idx_purpose = fp.find("时长运镜场景目的")
        if idx_purpose < 0:
            purpose_missing += 1
            continue
        idx_next = fp.find("时间分段叙事", idx_purpose)
        purpose_seg = fp[idx_purpose:idx_next] if idx_next > 0 else fp[idx_purpose:idx_purpose+300]
        if "此镜" not in purpose_seg and "戏剧目标" not in purpose_seg:
            purpose_missing += 1
    chk(27, "C5: Scene purpose sentence",
        purpose_missing <= 5,
        f"{purpose_missing} shots missing purpose")

    # C6: Ambient sound layer
    ambient_missing = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        if "环境音设计" not in fp or "环境音" not in fp:
            ambient_missing += 1
            continue
        idx_ambient = fp.find("环境音设计")
        idx_next = fp.find("负面提示词", idx_ambient)
        ambient_seg = fp[idx_ambient:idx_next] if idx_next > 0 else fp[idx_ambient:idx_ambient+200]
        if len(ambient_seg.strip()) < 15:
            ambient_missing += 1
    chk(28, "C6: Ambient sound layer",
        ambient_missing <= 5,
        f"{ambient_missing} shots missing ambient audio")

    # C7: Non-speaker mouth closed declaration
    nonspeaker_missing = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        sid = s.get("subshot_id", "?")
        # Check if shot has dialogue
        di = next((i for i in dp_items if i.get("subshot_id") == sid), {})
        diags = di.get("dialogue_refs", di.get("dialogue", []))
        has_dialogue = bool(diags) if isinstance(diags, list) else bool(diags)
        # For dialogue shots, non-speaking chars need mouth-closed declaration
        if has_dialogue:
            if "口型闭合" not in fp and "无口型" not in fp and "嘴部闭合" not in fp:
                nonspeaker_missing += 1
    chk(29, "C7: Non-speaker mouth closed",
        nonspeaker_missing <= 8,
        f"{nonspeaker_missing} dialogue shots missing mouth control")

    # C8: Word count 800-1800 (reinforce validate_composer_output check)
    wc_issues = 0
    for s in pkg.get("shots", pkg.get("items", [])):
        fp = s.get("full_prompt", "")
        if len(fp) < 800 or len(fp) > 1800:
            wc_issues += 1
    chk(30, "C8: Word count 800-1800",
        wc_issues <= 5,
        f"{wc_issues} shots outside range")

    total_checks = 30
    passed = total_checks - len(issues)
    print(f"\n  RESULT: {passed}/{total_checks} passed" + (f", {len(issues)} issues" if issues else " - ALL CLEAN"))
    if issues:
        print("  ISSUES:")
        for i, iss in enumerate(issues): print(f"    {i+1}. {iss}")
    return passed, len(issues)


if __name__ == "__main__":
    if QUALITY_MODE:
        args = [a for a in sys.argv if a != "--quality"]
        if len(args) >= 2:
            check_export("", args[1], quality_mode=True)
    elif len(sys.argv) >= 3:
        p, f = check_export(sys.argv[1], sys.argv[2])
        sys.exit(0 if f == 0 else 1)
    else:
        print("Usage: python check_export.py <export_md> <run_dir>")
        print("       python check_export.py --quality <run_dir>")
